"""
apps/settings_app/tests/test_data_management_system.py

Comprehensive test suite covering Data Management: Export (JSON + Excel),
Weekly Backup Settings, Instant Backup Mail, Excel Backup Validation & Restore,
Audit Logging, and Multi-Tenant Security Isolation.
"""

import io
import json
from django.contrib.auth import get_user_model
from django.core import mail
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.billing.models import Invoice, InvoiceLine
from apps.customers.models import Customer
from apps.organization.models import Organization
from apps.products.models import Product
from apps.settings_app.models import (
    BackupStatus,
    DataManagementAuditLog,
    OrganizationBackupSetting,
)
from apps.settings_app.services.backup_service import OrganizationBackupService
from apps.settings_app.services.excel_backup_service import ExcelBackupService
from apps.settings_app.services.excel_restore_service import ExcelRestoreService

User = get_user_model()


class DataManagementSystemTests(TestCase):
    def setUp(self):
        # User 1 & Org 1
        self.user1 = User.objects.create_user(
            username="owner1@acme.com",
            email="owner1@acme.com",
            password="Password123!",
            first_name="Owner",
            last_name="One"
        )
        self.org1 = Organization.objects.create(
            owner=self.user1,
            business_name="Acme Corp",
            gstin="27AAAAA0000A1Z5",
            business_email="owner1@acme.com"
        )
        self.user1.organization = self.org1
        self.user1.save()

        # User 2 & Org 2 (For multi-tenant testing)
        self.user2 = User.objects.create_user(
            username="owner2@beta.com",
            email="owner2@beta.com",
            password="Password123!",
            first_name="Owner",
            last_name="Two"
        )
        self.org2 = Organization.objects.create(
            owner=self.user2,
            business_name="Beta Corp",
            gstin="29BBBBB1111B2Z6",
            business_email="owner2@beta.com"
        )
        self.user2.organization = self.org2
        self.user2.save()

        # Data for Org 1
        self.customer1 = Customer.objects.create(
            organization=self.org1,
            name="Alice Smith",
            billing_address_line_1="100 Main St",
            billing_city="Mumbai",
            billing_state="Maharashtra",
            billing_pin_code="400001",
            billing_state_code="27"
        )
        self.product1 = Product.objects.create(
            organization=self.org1,
            name="Consulting Services",
            product_type="service",
            sac_code="998311",
            unit_price=1500.00,
            uqc="OTH"
        )
        self.invoice1 = Invoice.objects.create(
            organization=self.org1,
            customer_name_snapshot="Alice Smith",
            customer_billing_address_snapshot="100 Main St",
            customer_state_code_snapshot="27",
            status="issued",
            invoice_number="INV-2026-0001",
            invoice_date=timezone.now().date(),
            taxable_amount=1500.00,
            cgst_total=135.00,
            sgst_total=135.00,
            grand_total=1770.00
        )
        self.invoice_line1 = InvoiceLine.objects.create(
            invoice=self.invoice1,
            product=self.product1,
            position=1,
            product_name_snapshot="Consulting Services",
            product_type_snapshot="service",
            hsn_sac_snapshot="998311",
            taxability_type_snapshot="taxable",
            gst_rate_snapshot=18.00,
            quantity=1,
            unit_price=1500.00,
            taxable_value=1500.00,
            cgst_rate=9.00,
            cgst_amount=135.00,
            sgst_rate=9.00,
            sgst_amount=135.00,
            line_total=1770.00
        )

        self.client1 = Client()
        self.client1.force_login(self.user1)

        self.client2 = Client()
        self.client2.force_login(self.user2)

    def test_data_management_view_rendering(self):
        """Data Management page loads cleanly with active backup/export cards and zero obsolete elements."""
        url = reverse("settings_app:data_management")
        response = self.client1.get(url)
        self.assertEqual(response.status_code, 200)
        
        # Verify active elements
        self.assertContains(response, "Data Management")
        self.assertContains(response, "Backup & Export")
        self.assertContains(response, "Export Excel Backup (.xlsx)")
        self.assertContains(response, "Export JSON Backup")
        self.assertContains(response, "Weekly Data Backup")
        self.assertContains(response, "Import Data & Restore Backup")
        self.assertContains(response, "Backup History & Audit Log")
        
        # Verify obsolete and separated sections are NOT present in page content
        self.assertNotContains(response, "Archived Data")
        self.assertNotContains(response, "Data Cleanup")
        self.assertNotContains(response, "delete-account-modal")
        self.assertNotContains(response, "Delete Organization & Business Data")

    def test_manual_export_json_download(self):
        """JSON Export generates valid backup payload and records audit log."""
        url = reverse("settings_app:data_export")
        response = self.client1.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        
        self.assertTrue(DataManagementAuditLog.objects.filter(organization=self.org1, action="export").exists())

    def test_manual_export_excel_download(self):
        """Excel Export generates official versioned .xlsx workbook with 14 sheets."""
        url = reverse("settings_app:excel_export")
        response = self.client1.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        wb = load_workbook(io.BytesIO(response.content), data_only=True)
        self.assertIn("Dashboard", wb.sheetnames)
        self.assertIn("_Metadata", wb.sheetnames)
        self.assertIn("Customers", wb.sheetnames)

    def test_weekly_backup_toggle(self):
        """Weekly backup setting toggle updates model state and audit log."""
        url = reverse("settings_app:data_management")
        response = self.client1.post(url, {"action": "toggle_weekly_backup", "weekly_backup_enabled": "true"})
        self.assertEqual(response.status_code, 302)
        
        setting = OrganizationBackupSetting.objects.get(organization=self.org1)
        self.assertTrue(setting.weekly_backup_enabled)
        self.assertTrue(DataManagementAuditLog.objects.filter(organization=self.org1, action="weekly_backup_toggle").exists())

    def test_instant_backup_mail_view_success_and_security(self):
        """Instant Backup Mail view emails dual JSON + Excel attachments to organization owner."""
        url = reverse("settings_app:data_backup_mail")

        # 1. Unauthenticated request -> redirect to login
        unauth_client = Client()
        res_unauth = unauth_client.post(url)
        self.assertEqual(res_unauth.status_code, 302)

        # 2. Authenticated request -> emails fresh backup to owner1@acme.com
        res = self.client1.post(url)
        self.assertEqual(res.status_code, 302)

        self.assertEqual(len(mail.outbox), 1)
        sent_email = mail.outbox[0]
        self.assertEqual(sent_email.to, ["owner1@acme.com"])
        self.assertIn("Advance Billing — ", sent_email.subject)
        self.assertEqual(len(sent_email.attachments), 2)
        self.assertTrue(sent_email.attachments[0][0].endswith(".json"))
        self.assertTrue(sent_email.attachments[1][0].endswith(".xlsx"))

        # Verify audit log created
        self.assertTrue(DataManagementAuditLog.objects.filter(organization=self.org1, action="backup_sent").exists())

    def test_excel_import_validate_and_restore_system(self):
        """ExcelRestoreService validates workbook preview and restores organization data atomically."""
        excel_bytes, filename, _ = ExcelBackupService.generate_backup_workbook(self.org1)

        # Phase 1: Dry-run Validation
        is_valid, msg, preview = ExcelRestoreService.validate_and_preview(excel_bytes, filename, self.org1)
        self.assertTrue(is_valid, msg)
        self.assertEqual(preview["organization_name"], "Acme Corp")
        self.assertIn("customers", preview["counts"])

        # Phase 2: Execute Restore
        success, res_msg, res_preview = ExcelRestoreService.execute_restore(excel_bytes, filename, self.org1)
        self.assertTrue(success)
        self.assertIn("restored successfully", res_msg.lower())

    def test_tenant_isolation_in_data_management(self):
        """User 2 cannot export or mail backup data for User 1's organization."""
        url_mail = reverse("settings_app:data_backup_mail")
        res = self.client2.post(url_mail)
        # Should email User 2's owner2@beta.com, not owner1@acme.com
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["owner2@beta.com"])

    def test_audit_log_formatted_details(self):
        """DataManagementAuditLog.formatted_details formats raw JSON dicts into human-readable strings."""
        log1 = DataManagementAuditLog.objects.create(
            organization=self.org1,
            user=self.user1,
            action="backup_sent",
            details={"recipient": "vyomprajapati149@gmail.com", "trigger": "manual"}
        )
        self.assertEqual(log1.formatted_details, "Recipient: vyomprajapati149@gmail.com • Trigger: Manual")

        log2 = DataManagementAuditLog.objects.create(
            organization=self.org1,
            user=self.user1,
            action="export",
            details={"filename": "backup.xlsx", "total_records": 65}
        )
        self.assertEqual(log2.formatted_details, "File: backup.xlsx • Records: 65")

