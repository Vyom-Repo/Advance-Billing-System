"""
apps/settings_app/tests/test_weekly_backup.py

Comprehensive test suite for Weekly Organization Data Backup (JSON + Excel dual format) & Manual Backup system.
"""

import io
import json
import zipfile
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.core import mail
from django.core.management import call_command
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.billing.models import Invoice
from apps.customers.models import Customer
from apps.organization.models import Organization
from apps.products.models import Product
from apps.settings_app.models import (
    BackupStatus,
    BackupTrigger,
    OrganizationBackupLog,
    OrganizationBackupSetting,
)
from apps.settings_app.services.backup_service import OrganizationBackupService

User = get_user_model()


class WeeklyBackupAndExportTests(TestCase):
    def setUp(self):
        # User 1 & Org 1
        self.user1 = User.objects.create_user(
            username="owner1@acme.com",
            email="owner1@acme.com",
            password="Password123!",
            first_name="Owner",
            last_name="One"
        )
        self.org1 = Organization.objects.create(
            owner=self.user1,
            business_name="Acme Corp",
            gstin="27AAAAA0000A1Z5",
            business_email="owner1@acme.com",
            address_line_1="123 Tech Park",
            city="Mumbai",
            state="Maharashtra",
            pincode="400001"
        )
        self.user1.organization = self.org1
        self.user1.save()

        # User 2 & Org 2 (For multi-tenant security testing)
        self.user2 = User.objects.create_user(
            username="owner2@beta.com",
            email="owner2@beta.com",
            password="Password123!",
            first_name="Owner",
            last_name="Two"
        )
        self.org2 = Organization.objects.create(
            owner=self.user2,
            business_name="Beta Solutions",
            gstin="29BBBBB1111B2Z6",
            business_email="owner2@beta.com",
            address_line_1="456 Innovation Way",
            city="Bengaluru",
            state="Karnataka",
            pincode="560001"
        )
        self.user2.organization = self.org2
        self.user2.save()

        # Seed data for Org 1
        self.customer1 = Customer.objects.create(
            organization=self.org1,
            name="Alice Smith",
            billing_address_line_1="100 Main St",
            billing_city="Mumbai",
            billing_state="Maharashtra",
            billing_pin_code="400001",
            billing_state_code="27"
        )
        self.product1 = Product.objects.create(
            organization=self.org1,
            name="Consulting Services",
            product_type="service",
            sac_code="998311",
            unit_price=1500.00,
            uqc="OTH"
        )
        self.invoice1 = Invoice.objects.create(
            organization=self.org1,
            customer_name_snapshot="Alice Smith",
            customer_billing_address_snapshot="100 Main St",
            customer_state_code_snapshot="27",
            status="issued",
            invoice_number="INV-2026-0001",
            invoice_date=timezone.now().date(),
            grand_total=1770.00
        )

        # Clients
        self.client1 = Client()
        self.client1.force_login(self.user1)

        self.client2 = Client()
        self.client2.force_login(self.user2)

    def test_single_snapshot_generates_json_and_excel(self):
        """OrganizationBackupService.generate_single_snapshot creates matching JSON and Excel packages."""
        j_bytes, j_name, x_bytes, x_name, counts, total = OrganizationBackupService.generate_single_snapshot(self.org1)
        
        self.assertTrue(j_name.endswith(".json"))
        self.assertTrue(x_name.endswith(".xlsx"))
        self.assertGreater(len(j_bytes), 0)
        self.assertGreater(len(x_bytes), 0)

        # Validate JSON content
        j_data = json.loads(j_bytes.decode("utf-8"))
        self.assertEqual(j_data["metadata"]["signature"], "ADVANCE_BILLING_BACKUP")
        self.assertEqual(j_data["metadata"]["organization_id"], self.org1.id)
        self.assertEqual(len(j_data["customers"]), 1)
        self.assertEqual(j_data["customers"][0]["name"], "Alice Smith")

        # Validate Excel content
        wb = load_workbook(io.BytesIO(x_bytes), data_only=True)
        meta_ws = wb["_Metadata"]
        meta_dict = dict(row for row in meta_ws.iter_rows(values_only=True) if len(row) >= 2 and row[0])
        self.assertEqual(meta_dict["signature"], "ADVANCE_BILLING_BACKUP")
        self.assertEqual(str(meta_dict["organization_id"]), str(self.org1.id))

    def test_weekly_backup_scheduler_sends_dual_attachments(self):
        """Weekly backup scheduler generates both JSON + Excel attachments and emails them to owner."""
        setting = OrganizationBackupService.get_or_create_backup_setting(self.org1)
        setting.weekly_backup_enabled = True
        setting.save()

        call_command("run_weekly_backups", force=True)

        self.assertEqual(len(mail.outbox), 1)
        sent_email = mail.outbox[0]
        self.assertEqual(sent_email.to, ["owner1@acme.com"])
        self.assertIn("Advance Billing — Your Weekly Data Backup", sent_email.subject)
        
        # Dual attachments check: JSON + Excel
        self.assertEqual(len(sent_email.attachments), 2)
        
        att1_name, att1_bytes, att1_mime = sent_email.attachments[0]
        self.assertTrue(att1_name.endswith(".json"))
        self.assertEqual(att1_mime, "application/json")

        att2_name, att2_bytes, att2_mime = sent_email.attachments[1]
        self.assertTrue(att2_name.endswith(".xlsx"))
        self.assertEqual(att2_mime, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Verify setting & logs
        setting.refresh_from_db()
        self.assertEqual(setting.last_status, BackupStatus.SENT)
        log = OrganizationBackupLog.objects.filter(organization=self.org1, trigger=BackupTrigger.SCHEDULED).first()
        self.assertIsNotNone(log)
        self.assertEqual(log.status, BackupStatus.SENT)

    def test_manual_backup_mail_sends_dual_attachments(self):
        """Manual Backup Mail action sends both JSON + Excel attachments to organization owner."""
        url = reverse("settings_app:data_backup_mail")
        response = self.client1.post(url, HTTP_ACCEPT="application/json")
        self.assertEqual(response.status_code, 200)
        
        self.assertEqual(len(mail.outbox), 1)
        sent_email = mail.outbox[0]
        self.assertEqual(len(sent_email.attachments), 2)
        self.assertTrue(sent_email.attachments[0][0].endswith(".json"))
        self.assertTrue(sent_email.attachments[1][0].endswith(".xlsx"))

    def test_combined_attachment_size_limit(self):
        """Enforces 15 MB limit against total combined attachment size (JSON + Excel)."""
        setting = OrganizationBackupService.get_or_create_backup_setting(self.org1)
        setting.weekly_backup_enabled = True
        setting.save()

        fake_huge_bytes = b"X" * (16 * 1024 * 1024)  # 16 MB
        with patch.object(OrganizationBackupService, "generate_single_snapshot", return_value=(
            fake_huge_bytes, "backup.json", b"excel", "backup.xlsx", {}, 10
        )), patch.object(OrganizationBackupService, "validate_json_backup", return_value=(True, "OK")), \
           patch.object(OrganizationBackupService, "validate_excel_backup", return_value=(True, "OK")):
            success, msg = OrganizationBackupService.send_weekly_backup_email(self.org1, force=True)
            self.assertFalse(success)
            self.assertIn("exceed the 15 MB email limit", msg)
            
            setting.refresh_from_db()
            self.assertEqual(setting.last_status, BackupStatus.FAILED)

    def test_json_validation_rejects_unsupported_datasets(self):
        """JSON validation rejects packages containing unsupported datasets like 'payments'."""
        bad_json = {
            "metadata": {"signature": "ADVANCE_BILLING_BACKUP", "schema_version": "1.0", "organization_id": self.org1.id},
            "organization": {}, "customers": [], "customer_addresses": [], "products": [], "invoices": [], "invoice_items": [],
            "payments": [{"id": 1, "amount": 100}]
        }
        is_valid, msg = OrganizationBackupService.validate_json_backup(json.dumps(bad_json).encode("utf-8"), self.org1.id)
        self.assertFalse(is_valid)
        self.assertIn("Unsupported dataset 'payments'", msg)

    def test_tenant_isolation_in_dual_backup(self):
        """Dual backup contains ONLY Org 1 data and zero Org 2 data."""
        j_bytes, _, x_bytes, _, _, _ = OrganizationBackupService.generate_single_snapshot(self.org1)
        
        j_data = json.loads(j_bytes.decode("utf-8"))
        self.assertEqual(j_data["organization"]["business_name"], "Acme Corp")
        self.assertNotEqual(j_data["organization"]["business_name"], "Beta Solutions")

        wb = load_workbook(io.BytesIO(x_bytes), data_only=True)
        meta_ws = wb["_Metadata"]
        meta_dict = dict(row for row in meta_ws.iter_rows(values_only=True) if len(row) >= 2 and row[0])
        self.assertEqual(str(meta_dict["organization_id"]), str(self.org1.id))
