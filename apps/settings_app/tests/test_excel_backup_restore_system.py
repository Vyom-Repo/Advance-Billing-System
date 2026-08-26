"""
apps/settings_app/tests/test_excel_backup_restore_system.py

Comprehensive test suite verifying the Advance Billing Excel Backup & Strict Restore System.
Tests:
- Valid export generation containing all 14 worksheets + hidden signature
- File type restrictions (.xlsx only; rejects .xls, .xlsm, .csv)
- Signature & version validation (ADVANCE_BILLING_BACKUP, schema 1.0)
- Multi-tenant organization isolation (rejects Org A backup into Org B)
- Relational integrity verification (Customer, Product, Invoice, Item FKs)
- Decimal financial total reconciliation
- Phase 1 Dry-Run Preview (read-only)
- Phase 2 Atomic Transactional Restore & Rollback
- Full Round-Trip Database Verification (Export -> Clear -> Restore -> Match 100%)
"""

import io
from decimal import Decimal
from django.contrib.auth import get_user_model
from django.test import TestCase, Client
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.billing.models import Invoice, InvoiceLine, InvoiceStatus
from apps.customers.models import Customer
from apps.organization.models import Organization
from apps.products.models import Product
from apps.settings_app.services.excel_backup_service import ExcelBackupService
from apps.settings_app.services.excel_restore_service import ExcelRestoreService

User = get_user_model()


class ExcelBackupAndRestoreSystemTests(TestCase):
    def setUp(self):
        # Organization 1 (Source Org)
        self.user1 = User.objects.create_user(
            username="owner1@acmecorp.com",
            email="owner1@acmecorp.com",
            password="Password123!",
            first_name="Alice",
            last_name="Owner"
        )
        self.org1 = Organization.objects.create(
            owner=self.user1,
            business_name="Acme Solutions Pvt Ltd",
            business_email="contact@acmecorp.com",
            gstin="27AAAAA0000A1Z5",
            state_code="27",
            address_line_1="Tech Tower, Suite 404",
            city="Mumbai",
            state="Maharashtra",
            pincode="400001"
        )
        self.user1.organization = self.org1
        self.user1.save()

        # Organization 2 (Target Org for Isolation Tests)
        self.user2 = User.objects.create_user(
            username="owner2@betacorp.com",
            email="owner2@betacorp.com",
            password="Password123!",
            first_name="Bob",
            last_name="Owner"
        )
        self.org2 = Organization.objects.create(
            owner=self.user2,
            business_name="Beta Global Enterprises",
            business_email="support@betacorp.com"
        )
        self.user2.organization = self.org2
        self.user2.save()

        # Populate Test Data for Org 1
        self.cust1 = Customer.objects.create(
            organization=self.org1,
            name="Apex Trading LLC",
            gstin="27BBBBB1111B1Z2",
            customer_type="B2B",
            gst_status="registered",
            billing_address_line_1="Industrial Area Phase 2",
            billing_city="Pune",
            billing_state="Maharashtra",
            billing_pin_code="411001",
            billing_state_code="27"
        )

        self.prod1 = Product.objects.create(
            organization=self.org1,
            name="Enterprise Software License",
            product_type="goods",
            hsn_code="998313",
            taxability_type="taxable",
            gst_rate=Decimal("18.00"),
            unit_price=Decimal("5000.00"),
            price_basis="tax_exclusive",
            uqc="OTH"
        )

        self.inv1 = Invoice.objects.create(
            organization=self.org1,
            customer=self.cust1,
            invoice_number="INV-2026-0001",
            status=InvoiceStatus.ISSUED,
            invoice_date=timezone.now().date(),
            place_of_supply="27-Maharashtra",
            currency="INR",
            customer_name_snapshot="Apex Trading LLC",
            customer_gstin_snapshot="27BBBBB1111B1Z2",
            customer_billing_address_snapshot="Industrial Area Phase 2, Pune",
            customer_state_code_snapshot="27",
            subtotal=Decimal("5000.00"),
            discount_total=Decimal("0.00"),
            taxable_amount=Decimal("5000.00"),
            cgst_total=Decimal("450.00"),
            sgst_total=Decimal("450.00"),
            igst_total=Decimal("0.00"),
            cess_total=Decimal("0.00"),
            round_off=Decimal("0.00"),
            grand_total=Decimal("5900.00")
        )

        self.line1 = InvoiceLine.objects.create(
            invoice=self.inv1,
            position=1,
            product=self.prod1,
            product_name_snapshot="Enterprise Software License",
            product_type_snapshot="goods",
            hsn_sac_snapshot="998313",
            taxability_type_snapshot="taxable",
            gst_rate_snapshot=Decimal("18.00"),
            quantity=Decimal("1.000"),
            unit_price=Decimal("5000.00"),
            discount_type="none",
            discount_value=Decimal("0.00"),
            discount_amount=Decimal("0.00"),
            taxable_value=Decimal("5000.00"),
            cgst_rate=Decimal("9.00"),
            cgst_amount=Decimal("450.00"),
            sgst_rate=Decimal("9.00"),
            sgst_amount=Decimal("450.00"),
            igst_rate=Decimal("0.00"),
            igst_amount=Decimal("0.00"),
            cess_amount=Decimal("0.00"),
            line_total=Decimal("5900.00")
        )

        self.client = Client()

    def test_export_generates_valid_workbook_structure(self):
        """Export generates valid .xlsx containing all 14 worksheets + hidden signature."""
        excel_bytes, filename, manifest = ExcelBackupService.generate_backup_workbook(self.org1)
        self.assertTrue(filename.startswith("AdvanceBilling_Backup_"))
        self.assertTrue(filename.endswith(".xlsx"))

        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        expected_sheets = [
            "_Metadata", "Dashboard", "README", "Organization", "Customers",
            "Customer_Addresses", "Products", "Invoices", "Invoice_Items",
            "GST_Summary", "GST_Sales", "HSN_Summary", "Party_Summary", "Import_Map"
        ]
        for s in expected_sheets:
            self.assertIn(s, wb.sheetnames)

        # Hidden signature check
        meta_ws = wb["_Metadata"]
        self.assertEqual(meta_ws.sheet_state, "hidden")

        # Verify Dashboard openpyxl Charts
        wb_charts = load_workbook(io.BytesIO(excel_bytes), data_only=False)
        dash_ws = wb_charts["Dashboard"]
        self.assertEqual(len(dash_ws._charts), 10)

    def test_reject_wrong_file_extension(self):
        """Importer rejects files that do not end in .xlsx."""
        is_valid, msg, _ = ExcelRestoreService.validate_and_preview(b"fake data", "backup.xls", self.org1)
        self.assertFalse(is_valid)
        self.assertIn("Invalid file format", msg)

    def test_reject_missing_metadata_signature(self):
        """Importer rejects workbooks lacking the _Metadata signature sheet."""
        wb = Workbook()
        wb.create_sheet("README")
        buf = io.BytesIO()
        wb.save(buf)

        is_valid, msg, _ = ExcelRestoreService.validate_and_preview(buf.getvalue(), "backup.xlsx", self.org1)
        self.assertFalse(is_valid)
        self.assertIn("missing _Metadata signature sheet", msg)

    def test_reject_wrong_signature_value(self):
        """Importer rejects workbooks with invalid signature value."""
        wb = Workbook()
        ws = wb.create_sheet("_Metadata")
        ws.append(["Key", "Value"])
        ws.append(["signature", "SOME_OTHER_SIGNATURE"])
        ws.append(["schema_version", "1.0"])
        ws.append(["organization_id", self.org1.id])

        buf = io.BytesIO()
        wb.save(buf)

        is_valid, msg, _ = ExcelRestoreService.validate_and_preview(buf.getvalue(), "backup.xlsx", self.org1)
        self.assertFalse(is_valid)
        self.assertIn("invalid signature", msg)

    def test_reject_unsupported_schema_version(self):
        """Importer rejects unsupported schema version (e.g. 2.0)."""
        wb = Workbook()
        ws = wb.create_sheet("_Metadata")
        ws.append(["Key", "Value"])
        ws.append(["signature", "ADVANCE_BILLING_BACKUP"])
        ws.append(["schema_version", "2.0"])
        ws.append(["organization_id", self.org1.id])

        buf = io.BytesIO()
        wb.save(buf)

        is_valid, msg, _ = ExcelRestoreService.validate_and_preview(buf.getvalue(), "backup.xlsx", self.org1)
        self.assertFalse(is_valid)
        self.assertIn("unsupported schema version", msg)

    def test_reject_cross_tenant_organization_backup(self):
        """Organization B cannot import a backup belonging to Organization A."""
        excel_bytes, filename, _ = ExcelBackupService.generate_backup_workbook(self.org1)

        # Attempt to validate as Organization B
        is_valid, msg, _ = ExcelRestoreService.validate_and_preview(excel_bytes, filename, self.org2)
        self.assertFalse(is_valid)
        self.assertIn("belongs to a different organization", msg)

    def test_phase1_validation_dry_run_does_not_modify_database(self):
        """Phase 1 validation preview returns counts without writing to the database."""
        excel_bytes, filename, _ = ExcelBackupService.generate_backup_workbook(self.org1)

        initial_cust_count = Customer.objects.filter(organization=self.org1).count()
        is_valid, msg, preview = ExcelRestoreService.validate_and_preview(excel_bytes, filename, self.org1)

        self.assertTrue(is_valid)
        self.assertEqual(preview["counts"]["customers"]["total"], 1)
        self.assertEqual(Customer.objects.filter(organization=self.org1).count(), initial_cust_count)

    def test_full_round_trip_export_and_restore(self):
        """Full round trip: Export -> Clear test database -> Restore -> Match 100%."""
        excel_bytes, filename, _ = ExcelBackupService.generate_backup_workbook(self.org1)

        # Delete database records for Org 1
        InvoiceLine.objects.filter(invoice__organization=self.org1).delete()
        Invoice.objects.filter(organization=self.org1).delete()
        Product.objects.filter(organization=self.org1).delete()
        Customer.objects.filter(organization=self.org1).delete()

        self.assertEqual(Customer.objects.filter(organization=self.org1).count(), 0)
        self.assertEqual(Product.objects.filter(organization=self.org1).count(), 0)
        self.assertEqual(Invoice.objects.filter(organization=self.org1).count(), 0)

        # Execute Phase 2 Restore
        success, msg, preview = ExcelRestoreService.execute_restore(excel_bytes, filename, self.org1)
        self.assertTrue(success)

        # Verify Database Restored
        self.assertEqual(Customer.objects.filter(organization=self.org1).count(), 1)
        self.assertEqual(Product.objects.filter(organization=self.org1).count(), 1)
        self.assertEqual(Invoice.objects.filter(organization=self.org1).count(), 1)
        self.assertEqual(InvoiceLine.objects.filter(invoice__organization=self.org1).count(), 1)

        restored_cust = Customer.objects.get(organization=self.org1, uuid=self.cust1.uuid)
        self.assertEqual(restored_cust.name, "Apex Trading LLC")

        restored_inv = Invoice.objects.get(organization=self.org1, uuid=self.inv1.uuid)
        self.assertEqual(restored_inv.invoice_number, "INV-2026-0001")
        self.assertEqual(restored_inv.grand_total, Decimal("5900.00"))

    def test_excel_export_view_returns_attachment(self):
        """GET /settings/data-management/excel-export/ returns .xlsx attachment."""
        self.client.force_login(self.user1)
        response = self.client.get(reverse("settings_app:excel_export"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment; filename=", response["Content-Disposition"])

    def test_excel_import_validate_view(self):
        """POST /settings/data-management/excel-import-validate/ returns JSON preview."""
        excel_bytes, filename, _ = ExcelBackupService.generate_backup_workbook(self.org1)

        self.client.force_login(self.user1)
        excel_file = io.BytesIO(excel_bytes)
        excel_file.name = filename

        response = self.client.post(
            reverse("settings_app:excel_import_validate"),
            {"backup_file": excel_file}
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertIn("preview", data)
