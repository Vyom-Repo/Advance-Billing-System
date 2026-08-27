"""
apps/settings_app/services/backup_service.py

Service for generating structured organization data backups (JSON + Excel dual format)
and performing weekly scheduled email backup delivery to organization owners.
"""

import io
import json
import logging
import zipfile
from datetime import timedelta
from typing import Any, Dict, Tuple

from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import load_workbook

from apps.billing.models import Invoice, InvoiceLine
from apps.customers.models import Customer
from apps.organization.models import Organization
from apps.products.models import Product
from apps.settings_app.models import (
    BackupStatus,
    BackupTrigger,
    OrganizationBackupLog,
    OrganizationBackupSetting,
)

logger = logging.getLogger(__name__)

MAX_ATTACHMENT_SIZE_BYTES = 15 * 1024 * 1024  # 15 MB
MAX_EXPORT_RECORDS = 50_000  # Conservative server-side dataset boundary
SIGNATURE = "ADVANCE_BILLING_BACKUP"
SCHEMA_VERSION = "1.0"


class ExportDatasetTooLargeError(Exception):
    """Raised when an organization dataset exceeds the maximum allowed export limit."""

    pass


class OrganizationBackupService:
    @classmethod
    def count_organization_records(cls, organization: Organization) -> int:
        """
        Calculates total record count strictly scoped to the organization across
        customers, customer addresses, products, invoices, and invoice line items.
        Does NOT count unrelated organizations.
        """
        cust_count = Customer.objects.filter(organization=organization).count()
        prod_count = Product.objects.filter(organization=organization).count()
        inv_count = Invoice.objects.filter(organization=organization).count()
        item_count = InvoiceLine.objects.filter(invoice__organization=organization).count()
        return 1 + (cust_count * 2) + prod_count + inv_count + item_count

    @classmethod
    def get_or_create_backup_setting(cls, organization: Organization) -> OrganizationBackupSetting:
        setting, _ = OrganizationBackupSetting.objects.get_or_create(
            organization=organization,
            defaults={"weekly_backup_enabled": False, "last_status": BackupStatus.NEVER},
        )
        return setting

    @classmethod
    def serialize_datetime(cls, dt: Any) -> str:
        if dt is None:
            return ""
        if hasattr(dt, "isoformat"):
            return dt.isoformat()
        return str(dt)

    @classmethod
    def generate_single_snapshot(
        cls, organization: Organization
    ) -> Tuple[bytes, str, bytes, str, Dict[str, int], int]:
        """
        Retrieves all database data belonging strictly to the organization in ONE point-in-time snapshot.
        Generates BOTH:
        1. Complete machine-readable JSON backup
        2. Complete human-readable Excel backup (Dashboard + Reports + Data)

        Returns: (json_bytes, json_filename, excel_bytes, excel_filename, record_counts, total_records)
        """
        from apps.settings_app.services.excel_backup_service import ExcelBackupService  # noqa: PLC0415

        # Enforce server-side dataset safety boundary BEFORE loading full datasets
        total_records = cls.count_organization_records(organization)
        if total_records > MAX_EXPORT_RECORDS:
            err_msg = (
                f"Organization export dataset ({total_records} records) exceeds maximum allowed limit "
                f"of {MAX_EXPORT_RECORDS} records."
            )
            logger.warning(err_msg)
            raise ExportDatasetTooLargeError(err_msg)

        now = timezone.now()
        org_slug = (organization.business_name or "org").lower().replace(" ", "-")
        org_slug = "".join(c for c in org_slug if c.isalnum() or c in ("-", "_"))

        # 1. Organization Data
        org_dict = {
            "id": organization.id,
            "business_name": organization.business_name,
            "legal_business_name": organization.legal_business_name or "",
            "is_gst_registered": organization.is_gst_registered,
            "gstin": organization.gstin or "",
            "pan": organization.pan or "",
            "state_code": organization.state_code or "",
            "business_email": organization.business_email or "",
            "phone_number": organization.phone_number or "",
            "address_line_1": organization.address_line_1 or "",
            "address_line_2": organization.address_line_2 or "",
            "city": organization.city or "",
            "state": organization.state or "",
            "pincode": organization.pincode or "",
            "country": organization.country or "India",
            "signature_mode": organization.signature_mode or "none",
            "authorized_signatory_name": organization.authorized_signatory_name or "",
            "show_computer_generated_disclaimer": organization.show_computer_generated_disclaimer,
            "terms_and_conditions": organization.terms_and_conditions or "",
            "created_at": cls.serialize_datetime(organization.created_at),
            "updated_at": cls.serialize_datetime(organization.updated_at),
        }

        # 2. Customers & Customer Addresses
        customers_db = Customer.objects.filter(organization=organization)
        cust_list = []
        addr_list = []

        for c in customers_db:
            cid_str = str(c.uuid)
            cust_list.append({
                "customer_id": cid_str,
                "name": c.name,
                "gstin": c.gstin or "",
                "customer_type": c.customer_type,
                "gst_status": c.gst_status,
                "billing_address_line_1": c.billing_address_line_1 or "",
                "billing_address_line_2": c.billing_address_line_2 or "",
                "billing_city": c.billing_city or "",
                "billing_state": c.billing_state or "",
                "billing_pin_code": c.billing_pin_code or "",
                "billing_state_code": c.billing_state_code or "",
                "billing_country": c.billing_country or "India",
                "created_at": cls.serialize_datetime(c.created_at),
                "updated_at": cls.serialize_datetime(c.updated_at),
            })
            addr_list.append({
                "address_id": f"{cid_str}-addr",
                "customer_id": cid_str,
                "address_line_1": c.billing_address_line_1 or "",
                "address_line_2": c.billing_address_line_2 or "",
                "city": c.billing_city or "",
                "state": c.billing_state or "",
                "state_code": c.billing_state_code or "",
                "pincode": c.billing_pin_code or "",
                "country": c.billing_country or "India",
            })

        # 3. Products
        products_db = Product.objects.filter(organization=organization)
        prod_list = []
        for p in products_db:
            prod_list.append({
                "product_id": str(p.uuid),
                "name": p.name,
                "product_type": p.product_type,
                "hsn_code": p.hsn_code or "",
                "sac_code": p.sac_code or "",
                "taxability_type": p.taxability_type,
                "gst_rate": float(p.gst_rate) if p.gst_rate is not None else 0.0,
                "unit_price": float(p.unit_price) if p.unit_price is not None else 0.0,
                "price_basis": p.price_basis,
                "uqc": p.uqc,
                "cess_applicable": p.cess_applicable,
                "cess_type": p.cess_type or "",
                "cess_rate_or_amount": float(p.cess_rate_or_amount) if p.cess_rate_or_amount is not None else 0.0,
                "reverse_charge_applicable": p.reverse_charge_applicable,
                "created_at": cls.serialize_datetime(p.created_at),
                "updated_at": cls.serialize_datetime(p.updated_at),
            })

        # 4. Invoices
        invoices_db = Invoice.objects.filter(organization=organization).select_related("customer")
        inv_list = []
        for inv in invoices_db:
            inv_list.append({
                "invoice_id": str(inv.uuid),
                "invoice_number": inv.invoice_number,
                "status": inv.status,
                "invoice_date": cls.serialize_datetime(inv.invoice_date),
                "due_date": cls.serialize_datetime(inv.due_date),
                "place_of_supply": inv.place_of_supply or "",
                "currency": inv.currency or "INR",
                "customer_id": str(inv.customer.uuid) if inv.customer else "",
                "customer_name_snapshot": inv.customer_name_snapshot,
                "customer_gstin_snapshot": inv.customer_gstin_snapshot or "",
                "customer_billing_address_snapshot": inv.customer_billing_address_snapshot or "",
                "customer_state_code_snapshot": inv.customer_state_code_snapshot or "",
                "shipping_same_as_billing": inv.shipping_same_as_billing,
                "shipping_address_line_1": inv.shipping_address_line_1 or "",
                "shipping_city": inv.shipping_city or "",
                "shipping_state": inv.shipping_state or "",
                "shipping_pincode": inv.shipping_pincode or "",
                "subtotal": float(inv.subtotal),
                "discount_total": float(inv.discount_total),
                "taxable_amount": float(inv.taxable_amount),
                "cgst_total": float(inv.cgst_total),
                "sgst_total": float(inv.sgst_total),
                "igst_total": float(inv.igst_total),
                "cess_total": float(inv.cess_total),
                "round_off": float(inv.round_off),
                "grand_total": float(inv.grand_total),
                "notes": inv.notes or "",
                "terms": inv.terms or "",
                "created_at": cls.serialize_datetime(inv.created_at),
                "updated_at": cls.serialize_datetime(inv.updated_at),
            })

        # 5. Invoice Line Items
        items_db = InvoiceLine.objects.filter(invoice__organization=organization).select_related("invoice", "product")
        item_list = []
        for line in items_db:
            item_list.append({
                "item_id": line.id,
                "invoice_id": str(line.invoice.uuid),
                "product_id": str(line.product.uuid) if line.product else "",
                "position": line.position,
                "product_name_snapshot": line.product_name_snapshot,
                "product_type_snapshot": line.product_type_snapshot,
                "hsn_sac_snapshot": line.hsn_sac_snapshot or "",
                "taxability_type_snapshot": line.taxability_type_snapshot,
                "gst_rate_snapshot": float(line.gst_rate_snapshot),
                "cess_applicable_snapshot": line.cess_applicable_snapshot,
                "cess_type_snapshot": line.cess_type_snapshot or "",
                "cess_rate_snapshot": float(line.cess_rate_snapshot) if line.cess_rate_snapshot is not None else 0.0,
                "reverse_charge_snapshot": line.reverse_charge_snapshot,
                "price_basis_snapshot": line.price_basis_snapshot,
                "uqc_snapshot": line.uqc_snapshot,
                "quantity": float(line.quantity),
                "unit_price": float(line.unit_price),
                "discount_type": line.discount_type,
                "discount_value": float(line.discount_value),
                "discount_amount": float(line.discount_amount),
                "taxable_value": float(line.taxable_value),
                "cgst_rate": float(line.cgst_rate),
                "cgst_amount": float(line.cgst_amount),
                "sgst_rate": float(line.sgst_rate),
                "sgst_amount": float(line.sgst_amount),
                "igst_rate": float(line.igst_rate),
                "igst_amount": float(line.igst_amount),
                "cess_amount": float(line.cess_amount),
                "line_total": float(line.line_total),
            })

        record_counts = {
            "organization": 1,
            "customers": len(cust_list),
            "customer_addresses": len(addr_list),
            "products": len(prod_list),
            "invoices": len(inv_list),
            "invoice_items": len(item_list),
        }
        total_records = sum(record_counts.values())

        # Construct JSON Backup Payload
        json_payload = {
            "metadata": {
                "signature": SIGNATURE,
                "schema_version": SCHEMA_VERSION,
                "backup_type": "full",
                "organization_id": organization.id,
                "organization_name": organization.business_name,
                "created_at": now.isoformat(),
                "application": getattr(settings, "APP_NAME", "Advance Billing"),
            },
            "organization": org_dict,
            "customers": cust_list,
            "customer_addresses": addr_list,
            "products": prod_list,
            "invoices": inv_list,
            "invoice_items": item_list,
        }

        json_bytes = json.dumps(json_payload, indent=2).encode("utf-8")
        json_filename = f"AdvanceBilling_Backup_{org_slug}_{now.strftime('%Y-%m-%d')}.json"

        # Construct Excel Backup Payload from same DB state
        excel_bytes, excel_filename, _ = ExcelBackupService.generate_backup_workbook(organization)

        return json_bytes, json_filename, excel_bytes, excel_filename, record_counts, total_records

    @classmethod
    def generate_backup_datasets(cls, organization: Organization) -> Tuple[Dict[str, Any], Dict[str, int]]:
        """Backwards compatibility helper."""
        j_bytes, _, _, _, record_counts, _ = cls.generate_single_snapshot(organization)
        j_data = json.loads(j_bytes.decode("utf-8"))
        datasets = {
            "organization.json": [j_data["organization"]],
            "customers.json": j_data["customers"],
            "customer_addresses.json": j_data["customer_addresses"],
            "products.json": j_data["products"],
            "invoices.json": j_data["invoices"],
            "invoice_items.json": j_data["invoice_items"],
        }
        return datasets, record_counts

    @classmethod
    def generate_backup_zip(cls, organization: Organization) -> Tuple[bytes, str, Dict[str, Any]]:
        """Backwards compatibility helper for legacy ZIP backup export & restore."""
        j_bytes, j_name, x_bytes, x_name, record_counts, total_records = cls.generate_single_snapshot(organization)
        now = timezone.now()
        manifest = {
            "export_version": 1,
            "created_at": cls.serialize_datetime(now),
            "application": getattr(settings, "APP_NAME", "Advance Billing"),
            "organization": {
                "id": organization.id,
                "business_name": organization.business_name,
                "owner_email": organization.owner.email if organization.owner else "",
            },
            "datasets": record_counts,
            "total_records": total_records,
        }

        # Parse once cleanly to populate individual json files in zip
        j_data = json.loads(j_bytes)
        datasets = {
            "organization.json": json.dumps([j_data["organization"]], indent=2).encode("utf-8"),
            "customers.json": json.dumps(j_data["customers"], indent=2).encode("utf-8"),
            "customer_addresses.json": json.dumps(j_data["customer_addresses"], indent=2).encode("utf-8"),
            "products.json": json.dumps(j_data["products"], indent=2).encode("utf-8"),
            "invoices.json": json.dumps(j_data["invoices"], indent=2).encode("utf-8"),
            "invoice_items.json": json.dumps(j_data["invoice_items"], indent=2).encode("utf-8"),
        }
        del j_data

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("manifest.json", json.dumps(manifest, indent=2))
            for fn, content_bytes in datasets.items():
                zf.writestr(fn, content_bytes)
            zf.writestr(j_name, j_bytes)
            zf.writestr(x_name, x_bytes)

        zip_bytes = zip_buf.getvalue()
        zip_buf.close()
        filename = f"advance-billing-backup-{now.strftime('%Y-%m-%d')}.zip"
        return zip_bytes, filename, manifest

    @classmethod
    def validate_json_backup(cls, json_bytes: bytes, organization_id: int) -> Tuple[bool, str]:
        """Validates that json_bytes is a clean, well-formed Advance Billing JSON backup."""
        try:
            data = json.loads(json_bytes.decode("utf-8"))
            if not isinstance(data, dict):
                return False, "JSON root must be an object."
            meta = data.get("metadata", {})
            if meta.get("signature") != SIGNATURE:
                return False, f"Invalid JSON signature '{meta.get('signature')}'. Expected '{SIGNATURE}'."
            if str(meta.get("schema_version")) != SCHEMA_VERSION:
                return False, f"Unsupported JSON schema version '{meta.get('schema_version')}'."
            if meta.get("organization_id") != organization_id:
                return False, "JSON organization_id mismatch."

            req_keys = ["metadata", "organization", "customers", "customer_addresses", "products", "invoices", "invoice_items"]
            for k in req_keys:
                if k not in data:
                    return False, f"Missing required key '{k}' in backup JSON."

            unsupported = ["payments", "expenses", "notifications", "settings", "receivables", "payment_collections"]
            for un in unsupported:
                if un in data:
                    return False, f"Unsupported dataset '{un}' present in backup JSON."
            return True, "Valid JSON Backup"
        except Exception as e:
            return False, f"Corrupt JSON backup payload: {str(e)}"

    @classmethod
    def validate_excel_backup(cls, excel_bytes: bytes, organization_id: int) -> Tuple[bool, str]:
        """Validates that excel_bytes is a clean, well-formed Advance Billing Excel backup."""
        try:
            wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
            if "_Metadata" not in wb.sheetnames:
                return False, "Missing _Metadata sheet in Excel backup."
            meta_ws = wb["_Metadata"]
            meta_dict = {}
            for row in meta_ws.iter_rows(values_only=True):
                if len(row) >= 2 and row[0]:
                    meta_dict[str(row[0])] = row[1]

            if meta_dict.get("signature") != SIGNATURE:
                return False, f"Invalid signature '{meta_dict.get('signature')}' in Excel _Metadata."
            if str(meta_dict.get("schema_version")) != SCHEMA_VERSION:
                return False, f"Unsupported schema version '{meta_dict.get('schema_version')}' in Excel."
            if str(meta_dict.get("organization_id")) != str(organization_id):
                return False, "Organization ID mismatch in Excel backup."

            required_sheets = ["Dashboard", "README", "Organization", "Customers", "Customer_Addresses", "Products", "Invoices", "Invoice_Items"]
            for req in required_sheets:
                if req not in wb.sheetnames:
                    return False, f"Missing required sheet '{req}' in Excel backup."
            return True, "Valid Excel Backup"
        except Exception as e:
            return False, f"Corrupt Excel backup payload: {str(e)}"

    @classmethod
    def send_weekly_backup_email(
        cls, organization: Organization, force: bool = False, trigger: str = BackupTrigger.SCHEDULED
    ) -> Tuple[bool, str]:
        """
        Generates and emails BOTH JSON + Excel backup attachments to the organization owner.
        - Enforces idempotency (6 days) for scheduled runs unless force=True.
        - Validates both attachments.
        - Enforces 15 MB payload limit against combined (JSON + Excel) file size.
        - Cleans up memory safely in a finally block.
        """
        owner = getattr(organization, "owner", None)
        recipient_email = owner.email if owner else getattr(organization, "business_email", "")
        if not recipient_email:
            err_msg = f"Organization {organization.business_name} owner has no valid email address."
            logger.error(err_msg)
            return False, err_msg

        setting = cls.get_or_create_backup_setting(organization)
        now = timezone.now()

        # Idempotency check: Skip if backup sent in the last 6 days unless forced
        if not force and setting.last_backup_at:
            if (now - setting.last_backup_at) < timedelta(days=6):
                msg = f"Weekly backup already sent for {organization.business_name} on {setting.last_backup_at.strftime('%Y-%m-%d')}."
                logger.info(msg)
                return True, msg

        # Check dataset limit before generation
        total_records = cls.count_organization_records(organization)
        if total_records > MAX_EXPORT_RECORDS:
            err_msg = (
                f"Organization dataset ({total_records} records) exceeds maximum allowed limit "
                f"of {MAX_EXPORT_RECORDS} records."
            )
            logger.warning(err_msg)
            raise ExportDatasetTooLargeError(err_msg)

        setting.last_status = BackupStatus.GENERATING
        setting.save(update_fields=["last_status", "updated_at"])

        json_bytes = None
        excel_bytes = None

        try:
            # 1. Generate single-snapshot dual backup (JSON + Excel)
            (
                json_bytes,
                json_filename,
                excel_bytes,
                excel_filename,
                record_counts,
                total_records,
            ) = cls.generate_single_snapshot(organization)

            # 2. Validate JSON payload
            j_valid, j_err = cls.validate_json_backup(json_bytes, organization.id)
            if not j_valid:
                raise ValueError(f"JSON validation failed: {j_err}")

            # 3. Validate Excel payload
            x_valid, x_err = cls.validate_excel_backup(excel_bytes, organization.id)
            if not x_valid:
                raise ValueError(f"Excel validation failed: {x_err}")

            # 4. Total attachment size check (JSON + Excel)
            combined_size = len(json_bytes) + len(excel_bytes)
            if combined_size > MAX_ATTACHMENT_SIZE_BYTES:
                err_msg = f"Weekly backup attachments ({combined_size / (1024*1024):.2f} MB) exceed the 15 MB email limit."
                logger.error(err_msg)

                setting.last_status = BackupStatus.FAILED
                setting.last_error = err_msg
                setting.save(update_fields=["last_status", "last_error", "updated_at"])

                OrganizationBackupLog.objects.create(
                    organization=organization,
                    trigger=trigger,
                    status=BackupStatus.FAILED,
                    record_count=total_records,
                    file_size_bytes=combined_size,
                    error_message=err_msg,
                    recipient_email=recipient_email,
                )
                return False, err_msg

            # 5. Prepare Email Content with Advance Billing Branding ONLY
            from apps.common.services.email_service import AdvanceBillingEmailBranding  # noqa: PLC0415
            from django.utils.html import strip_tags  # noqa: PLC0415

            email_branding = AdvanceBillingEmailBranding.get_email_branding()
            org_name = organization.business_name if organization else "Advance Billing"
            owner_name = owner.get_full_name() or owner.first_name or owner.email.split("@")[0].capitalize() if owner else "Valued Customer"

            context = {
                "email_branding": email_branding,
                "branding": email_branding,
                "org_name": org_name,
                "organization": organization,
                "owner_name": owner_name,
                "backup_date": now.strftime("%d %b %Y"),
                "total_records": total_records,
                "counts": record_counts,
                "json_filename": json_filename,
                "excel_filename": excel_filename,
                "total_size_mb": f"{combined_size / (1024*1024):.2f}",
            }

            subject = f"Advance Billing — Your Weekly Data Backup ({org_name})"

            html_content = render_to_string("emails/backup_email.html", context)
            body_text = strip_tags(html_content)

            email = EmailMultiAlternatives(
                subject=subject,
                body=body_text,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[recipient_email],
            )
            email.attach_alternative(html_content, "text/html")
            email.attach(json_filename, json_bytes, "application/json")
            email.attach(excel_filename, excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            email.send(fail_silently=False)

            # 6. Record Success in Audit Log & Settings
            setting.last_backup_at = now
            if trigger == BackupTrigger.SCHEDULED or not setting.next_backup_at:
                setting.next_backup_at = now + timedelta(days=7)
            setting.last_status = BackupStatus.SENT
            setting.last_error = ""
            setting.last_record_count = total_records
            setting.save(update_fields=[
                "last_backup_at", "next_backup_at", "last_status",
                "last_error", "last_record_count", "updated_at"
            ])

            OrganizationBackupLog.objects.create(
                organization=organization,
                trigger=trigger,
                status=BackupStatus.SENT,
                record_count=total_records,
                file_size_bytes=combined_size,
                recipient_email=recipient_email,
            )

            success_msg = f"Weekly backup (JSON + Excel) successfully emailed to {recipient_email} for {organization.business_name}."
            logger.info(success_msg)
            return True, success_msg

        except Exception as e:
            err_msg = f"Failed to send backup email for {organization.business_name}: {str(e)}"
            logger.exception(err_msg)

            setting.last_status = BackupStatus.FAILED
            setting.last_error = str(e)
            setting.save(update_fields=["last_status", "last_error", "updated_at"])

            OrganizationBackupLog.objects.create(
                organization=organization,
                trigger=trigger,
                status=BackupStatus.FAILED,
                error_message=str(e),
                recipient_email=recipient_email,
            )
            return False, err_msg

        finally:
            # Release memory safely
            if json_bytes is not None:
                del json_bytes
            if excel_bytes is not None:
                del excel_bytes
