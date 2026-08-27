"""
apps/settings_app/services/excel_backup_service.py

Service for exporting complete, versioned, signed Advance Billing Excel Backup workbooks (.xlsx).
Follows the approved 14-worksheet specification with a visually polished Business Dashboard:
- Dashboard (with Excel formulas, 8 KPI cards, and 10 openpyxl charts)
- README (human-readable backup metadata)
- Organization, Customers, Customer_Addresses, Products, Invoices, Invoice_Items (6 normalized source sheets)
- GST_Summary, GST_Sales, HSN_Summary, Party_Summary (analytical reporting sheets)
- Import_Map (schema column mapping metadata)
- _Metadata (hidden machine-readable identity signature)
"""

import io
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from typing import Any, Dict, Tuple

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from apps.billing.models import Invoice, InvoiceLine
from apps.customers.models import Customer
from apps.organization.models import Organization
from apps.products.models import Product

SIGNATURE = "ADVANCE_BILLING_BACKUP"
SCHEMA_VERSION = "1.0"


class ExcelBackupService:
    @classmethod
    def serialize_val(cls, val: Any) -> Any:
        if val is None:
            return ""
        if isinstance(val, Decimal):
            return float(val)
        if hasattr(val, "isoformat"):
            return val.isoformat()
        return str(val)

    @classmethod
    def apply_header_styles(cls, ws, num_cols: int, title: str = ""):
        ws.views.sheetView[0].showGridLines = True
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

    @classmethod
    def auto_fit_columns(cls, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    @classmethod
    def generate_backup_workbook(cls, organization: Organization) -> Tuple[bytes, str, Dict[str, Any]]:
        from apps.settings_app.services.backup_service import (  # noqa: PLC0415
            OrganizationBackupService,
            MAX_EXPORT_RECORDS,
            ExportDatasetTooLargeError,
        )

        total_records = OrganizationBackupService.count_organization_records(organization)
        if total_records > MAX_EXPORT_RECORDS:
            err_msg = (
                f"Organization dataset ({total_records} records) exceeds maximum allowed Excel export limit "
                f"of {MAX_EXPORT_RECORDS} records."
            )
            raise ExportDatasetTooLargeError(err_msg)

        wb = Workbook()
        default_ws = wb.active

        now = timezone.now()
        org_slug = (organization.business_name or "org").lower().replace(" ", "-")
        org_slug = "".join(c for c in org_slug if c.isalnum() or c in ("-", "_"))
        filename = f"AdvanceBilling_Backup_{org_slug}_{now.strftime('%Y-%m-%d')}.xlsx"

        # ---------------------------------------------------------
        # 1. Hidden _Metadata Sheet
        # ---------------------------------------------------------
        ws_meta = wb.create_sheet(title="_Metadata")
        ws_meta.views.sheetView[0].showGridLines = False
        ws_meta.append(["Key", "Value"])
        ws_meta.append(["signature", SIGNATURE])
        ws_meta.append(["schema_version", SCHEMA_VERSION])
        ws_meta.append(["application", getattr(settings, "APP_NAME", "Advance Billing")])
        ws_meta.append(["organization_id", organization.id])
        ws_meta.append(["organization_name", organization.business_name])
        ws_meta.append(["created_at", now.isoformat()])
        ws_meta.sheet_state = "hidden"

        # ---------------------------------------------------------
        # 2. README Sheet
        # ---------------------------------------------------------
        ws_readme = wb.create_sheet(title="README")
        ws_readme.views.sheetView[0].showGridLines = True
        ws_readme.column_dimensions["A"].width = 28
        ws_readme.column_dimensions["B"].width = 50

        title_cell = ws_readme.cell(row=1, column=1, value="ADVANCE BILLING — OFFICIAL SYSTEM BACKUP")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FF7A00")
        ws_readme.merge_cells("A1:B1")

        readme_data = [
            ("Backup Signature", SIGNATURE),
            ("Export Type", "Full Organization Backup + Dashboard"),
            ("Application", getattr(settings, "APP_NAME", "Advance Billing")),
            ("Schema Version", SCHEMA_VERSION),
            ("Organization ID", str(organization.id)),
            ("Organization Name", organization.business_name),
            ("Export Date", now.strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Currency", "INR (₹)"),
            ("Restoration Policy", "Restores normalized transactional sheets only. Reporting sheets are regenerated post-restore."),
        ]

        for idx, (label, val) in enumerate(readme_data, start=3):
            cell_lbl = ws_readme.cell(row=idx, column=1, value=label)
            cell_lbl.font = Font(name="Calibri", size=11, bold=True)
            cell_val = ws_readme.cell(row=idx, column=2, value=val)
            cell_val.font = Font(name="Calibri", size=11)

        # ---------------------------------------------------------
        # 3. Normalized Data Collection
        # ---------------------------------------------------------
        # Organization
        org_rows = [[
            organization.id,
            organization.business_name,
            organization.legal_business_name or "",
            organization.is_gst_registered,
            organization.gstin or "",
            organization.pan or "",
            organization.state_code or "",
            organization.business_email or "",
            organization.phone_number or "",
            organization.address_line_1 or "",
            organization.address_line_2 or "",
            organization.city or "",
            organization.state or "",
            organization.pincode or "",
            organization.country or "India",
            organization.signature_mode or "none",
            organization.authorized_signatory_name or "",
            organization.show_computer_generated_disclaimer,
            organization.terms_and_conditions or "",
            cls.serialize_val(organization.created_at),
            cls.serialize_val(organization.updated_at),
        ]]

        # Customers & Customer Addresses
        customers_db = Customer.objects.filter(organization=organization)
        cust_rows = []
        addr_rows = []

        for c in customers_db:
            cid_str = str(c.uuid)
            cust_rows.append([
                cid_str,
                c.name,
                c.gstin or "",
                c.customer_type,
                c.gst_status,
                c.billing_address_line_1 or "",
                c.billing_address_line_2 or "",
                c.billing_city or "",
                c.billing_state or "",
                c.billing_pin_code or "",
                c.billing_state_code or "",
                c.billing_country or "India",
                cls.serialize_val(c.created_at),
                cls.serialize_val(c.updated_at),
            ])
            addr_rows.append([
                f"{cid_str}-addr",
                cid_str,
                c.billing_address_line_1 or "",
                c.billing_address_line_2 or "",
                c.billing_city or "",
                c.billing_state or "",
                c.billing_state_code or "",
                c.billing_pin_code or "",
                c.billing_country or "India",
            ])

        # Products
        products_db = Product.objects.filter(organization=organization)
        prod_rows = []
        for p in products_db:
            prod_rows.append([
                str(p.uuid),
                p.name,
                p.product_type,
                p.hsn_code or "",
                p.sac_code or "",
                p.taxability_type,
                float(p.gst_rate) if p.gst_rate is not None else 0.0,
                float(p.unit_price) if p.unit_price is not None else 0.0,
                p.price_basis,
                p.uqc,
                p.cess_applicable,
                p.cess_type or "",
                float(p.cess_rate_or_amount) if p.cess_rate_or_amount is not None else 0.0,
                p.reverse_charge_applicable,
                cls.serialize_val(p.created_at),
                cls.serialize_val(p.updated_at),
            ])

        # Invoices
        invoices_db = Invoice.objects.filter(organization=organization).select_related("customer")
        inv_rows = []
        for inv in invoices_db:
            inv_rows.append([
                str(inv.uuid),
                inv.invoice_number,
                inv.status,
                cls.serialize_val(inv.invoice_date),
                cls.serialize_val(inv.due_date),
                inv.place_of_supply or "",
                inv.currency or "INR",
                str(inv.customer.uuid) if inv.customer else "",
                inv.customer_name_snapshot,
                inv.customer_gstin_snapshot or "",
                inv.customer_billing_address_snapshot or "",
                inv.customer_state_code_snapshot or "",
                inv.shipping_same_as_billing,
                inv.shipping_address_line_1 or "",
                inv.shipping_city or "",
                inv.shipping_state or "",
                inv.shipping_pincode or "",
                float(inv.subtotal),
                float(inv.discount_total),
                float(inv.taxable_amount),
                float(inv.cgst_total),
                float(inv.sgst_total),
                float(inv.igst_total),
                float(inv.cess_total),
                float(inv.round_off),
                float(inv.grand_total),
                inv.notes or "",
                inv.terms or "",
                cls.serialize_val(inv.created_at),
                cls.serialize_val(inv.updated_at),
            ])

        # Invoice Items
        items_db = InvoiceLine.objects.filter(invoice__organization=organization).select_related("invoice", "product")
        item_rows = []
        for line in items_db:
            item_rows.append([
                line.id,
                str(line.invoice.uuid),
                str(line.product.uuid) if line.product else "",
                line.position,
                line.product_name_snapshot,
                line.product_type_snapshot,
                line.hsn_sac_snapshot or "",
                line.taxability_type_snapshot,
                float(line.gst_rate_snapshot),
                line.cess_applicable_snapshot,
                line.cess_type_snapshot or "",
                float(line.cess_rate_snapshot) if line.cess_rate_snapshot is not None else 0.0,
                line.reverse_charge_snapshot,
                line.price_basis_snapshot,
                line.uqc_snapshot,
                float(line.quantity),
                float(line.unit_price),
                line.discount_type,
                float(line.discount_value),
                float(line.discount_amount),
                float(line.taxable_value),
                float(line.cgst_rate),
                float(line.cgst_amount),
                float(line.sgst_rate),
                float(line.sgst_amount),
                float(line.igst_rate),
                float(line.igst_amount),
                float(line.cess_amount),
                float(line.line_total),
            ])

        # ---------------------------------------------------------
        # 4. Write Normalized Sheets
        # ---------------------------------------------------------
        headers_dict = {
            "Organization": [
                "organization_id", "business_name", "legal_business_name", "is_gst_registered",
                "gstin", "pan", "state_code", "business_email", "phone_number", "address_line_1",
                "address_line_2", "city", "state", "pincode", "country", "signature_mode",
                "authorized_signatory_name", "show_computer_generated_disclaimer",
                "terms_and_conditions", "created_at", "updated_at"
            ],
            "Customers": [
                "customer_id", "name", "gstin", "customer_type", "gst_status",
                "billing_address_line_1", "billing_address_line_2", "billing_city",
                "billing_state", "billing_pin_code", "billing_state_code", "billing_country",
                "created_at", "updated_at"
            ],
            "Customer_Addresses": [
                "address_id", "customer_id", "address_line_1", "address_line_2",
                "city", "state", "state_code", "pincode", "country"
            ],
            "Products": [
                "product_id", "name", "product_type", "hsn_code", "sac_code",
                "taxability_type", "gst_rate", "unit_price", "price_basis", "uqc",
                "cess_applicable", "cess_type", "cess_rate_or_amount",
                "reverse_charge_applicable", "created_at", "updated_at"
            ],
            "Invoices": [
                "invoice_id", "invoice_number", "status", "invoice_date", "due_date",
                "place_of_supply", "currency", "customer_id", "customer_name_snapshot",
                "customer_gstin_snapshot", "customer_billing_address_snapshot",
                "customer_state_code_snapshot", "shipping_same_as_billing",
                "shipping_address_line_1", "shipping_city", "shipping_state",
                "shipping_pincode", "subtotal", "discount_total", "taxable_amount",
                "cgst_total", "sgst_total", "igst_total", "cess_total", "round_off",
                "grand_total", "notes", "terms", "created_at", "updated_at"
            ],
            "Invoice_Items": [
                "item_id", "invoice_id", "product_id", "position", "product_name_snapshot",
                "product_type_snapshot", "hsn_sac_snapshot", "taxability_type_snapshot",
                "gst_rate_snapshot", "cess_applicable_snapshot", "cess_type_snapshot",
                "cess_rate_snapshot", "reverse_charge_snapshot", "price_basis_snapshot",
                "uqc_snapshot", "quantity", "unit_price", "discount_type", "discount_value",
                "discount_amount", "taxable_value", "cgst_rate", "cgst_amount", "sgst_rate",
                "sgst_amount", "igst_rate", "igst_amount", "cess_amount", "line_total"
            ],
        }

        data_rows_dict = {
            "Organization": org_rows,
            "Customers": cust_rows,
            "Customer_Addresses": addr_rows,
            "Products": prod_rows,
            "Invoices": inv_rows,
            "Invoice_Items": item_rows,
        }

        for sheet_name in ["Organization", "Customers", "Customer_Addresses", "Products", "Invoices", "Invoice_Items"]:
            ws = wb.create_sheet(title=sheet_name)
            headers = headers_dict[sheet_name]
            ws.append(headers)
            cls.apply_header_styles(ws, len(headers))
            for row in data_rows_dict[sheet_name]:
                ws.append(row)
            cls.auto_fit_columns(ws)

        # ---------------------------------------------------------
        # 5. Reporting & Analytical Sheets
        # ---------------------------------------------------------
        # GST Summary
        ws_gst_sum = wb.create_sheet(title="GST_Summary")
        gst_sum_headers = ["GST Rate (%)", "Taxable Amount (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Cess (₹)", "Total GST Tax (₹)"]
        ws_gst_sum.append(gst_sum_headers)
        cls.apply_header_styles(ws_gst_sum, len(gst_sum_headers))

        gst_rates_map = {}
        for line in items_db:
            rate = float(line.gst_rate_snapshot)
            if rate not in gst_rates_map:
                gst_rates_map[rate] = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0}
            gst_rates_map[rate]["taxable"] += float(line.taxable_value)
            gst_rates_map[rate]["cgst"] += float(line.cgst_amount)
            gst_rates_map[rate]["sgst"] += float(line.sgst_amount)
            gst_rates_map[rate]["igst"] += float(line.igst_amount)
            gst_rates_map[rate]["cess"] += float(line.cess_amount)

        for rate in sorted(gst_rates_map.keys()):
            d = gst_rates_map[rate]
            tot_tax = d["cgst"] + d["sgst"] + d["igst"] + d["cess"]
            ws_gst_sum.append([rate, d["taxable"], d["cgst"], d["sgst"], d["igst"], d["cess"], tot_tax])
        cls.auto_fit_columns(ws_gst_sum)

        # GST Sales
        ws_gst_sales = wb.create_sheet(title="GST_Sales")
        gst_sales_headers = ["Invoice Number", "Invoice Date", "Customer Name", "Customer GSTIN", "POS", "Invoice Type", "Taxable Amount", "Total Tax", "Grand Total"]
        ws_gst_sales.append(gst_sales_headers)
        cls.apply_header_styles(ws_gst_sales, len(gst_sales_headers))

        for inv in invoices_db:
            inv_type = "B2B" if inv.customer_gstin_snapshot else "B2C"
            tot_tax = float(inv.cgst_total + inv.sgst_total + inv.igst_total + inv.cess_total)
            ws_gst_sales.append([
                inv.invoice_number,
                cls.serialize_val(inv.invoice_date),
                inv.customer_name_snapshot,
                inv.customer_gstin_snapshot or "URP",
                inv.place_of_supply or "",
                inv_type,
                float(inv.taxable_amount),
                tot_tax,
                float(inv.grand_total)
            ])
        cls.auto_fit_columns(ws_gst_sales)

        # HSN Summary
        ws_hsn = wb.create_sheet(title="HSN_Summary")
        hsn_headers = ["HSN/SAC Code", "Description", "Total Quantity", "UQC", "Taxable Value", "Total Tax Amount"]
        ws_hsn.append(hsn_headers)
        cls.apply_header_styles(ws_hsn, len(hsn_headers))

        hsn_map = {}
        for line in items_db:
            hsn = line.hsn_sac_snapshot or "N/A"
            if hsn not in hsn_map:
                hsn_map[hsn] = {"desc": line.product_name_snapshot, "qty": 0.0, "uqc": line.uqc_snapshot, "taxable": 0.0, "tax": 0.0}
            hsn_map[hsn]["qty"] += float(line.quantity)
            hsn_map[hsn]["taxable"] += float(line.taxable_value)
            hsn_map[hsn]["tax"] += float(line.cgst_amount + line.sgst_amount + line.igst_amount + line.cess_amount)

        for hsn, d in hsn_map.items():
            ws_hsn.append([hsn, d["desc"], d["qty"], d["uqc"], d["taxable"], d["tax"]])
        cls.auto_fit_columns(ws_hsn)

        # Party Summary
        ws_party = wb.create_sheet(title="Party_Summary")
        party_headers = ["Customer Name", "GSTIN", "Total Invoices", "Total Taxable Value", "Total GST Tax", "Total Invoice Value"]
        ws_party.append(party_headers)
        cls.apply_header_styles(ws_party, len(party_headers))

        party_map = {}
        for inv in invoices_db:
            cname = inv.customer_name_snapshot
            if cname not in party_map:
                party_map[cname] = {"gstin": inv.customer_gstin_snapshot or "URP", "count": 0, "taxable": 0.0, "tax": 0.0, "grand": 0.0}
            party_map[cname]["count"] += 1
            party_map[cname]["taxable"] += float(inv.taxable_amount)
            party_map[cname]["tax"] += float(inv.cgst_total + inv.sgst_total + inv.igst_total + inv.cess_total)
            party_map[cname]["grand"] += float(inv.grand_total)

        for cname, d in party_map.items():
            ws_party.append([cname, d["gstin"], d["count"], d["taxable"], d["tax"], d["grand"]])
        cls.auto_fit_columns(ws_party)

        # Import Map
        ws_map = wb.create_sheet(title="Import_Map")
        ws_map.append(["Sheet", "Key Column", "Restore Target Model", "Description"])
        cls.apply_header_styles(ws_map, 4)
        map_info = [
            ("Organization", "organization_id", "apps.organization.models.Organization", "Target Organization Details"),
            ("Customers", "customer_id", "apps.customers.models.Customer", "Customer Master Records"),
            ("Customer_Addresses", "address_id", "apps.customers.models.CustomerAddress", "Customer Billing Addresses"),
            ("Products", "product_id", "apps.products.models.Product", "Product & Service Master"),
            ("Invoices", "invoice_id", "apps.billing.models.Invoice", "Invoice Headers"),
            ("Invoice_Items", "item_id", "apps.billing.models.InvoiceLine", "Invoice Line Items"),
        ]
        for row in map_info:
            ws_map.append(list(row))
        cls.auto_fit_columns(ws_map)

        # ---------------------------------------------------------
        # 6. Chart Supporting Data Section (_ChartData)
        # ---------------------------------------------------------
        ws_chart_data = wb.create_sheet(title="_ChartData")
        ws_chart_data.views.sheetView[0].showGridLines = False

        # 1. Monthly Aggregations
        monthly_map = defaultdict(lambda: {"sales": 0.0, "count": 0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "taxable": 0.0})
        for inv in invoices_db:
            m_key = inv.invoice_date.strftime("%b-%Y") if hasattr(inv.invoice_date, "strftime") else str(inv.invoice_date)[:7]
            monthly_map[m_key]["sales"] += float(inv.grand_total)
            monthly_map[m_key]["count"] += 1
            monthly_map[m_key]["cgst"] += float(inv.cgst_total)
            monthly_map[m_key]["sgst"] += float(inv.sgst_total)
            monthly_map[m_key]["igst"] += float(inv.igst_total)
            monthly_map[m_key]["taxable"] += float(inv.taxable_amount)

        if not monthly_map:
            monthly_map["No Data"] = {"sales": 0.0, "count": 0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "taxable": 0.0}

        ws_chart_data.cell(row=1, column=1, value="Month")
        ws_chart_data.cell(row=1, column=2, value="Sales (₹)")
        ws_chart_data.cell(row=1, column=3, value="Invoices")
        ws_chart_data.cell(row=1, column=4, value="CGST (₹)")
        ws_chart_data.cell(row=1, column=5, value="SGST (₹)")
        ws_chart_data.cell(row=1, column=6, value="IGST (₹)")
        ws_chart_data.cell(row=1, column=7, value="Taxable Value (₹)")

        m_idx = 2
        for m_lbl, m_d in monthly_map.items():
            ws_chart_data.cell(row=m_idx, column=1, value=m_lbl)
            ws_chart_data.cell(row=m_idx, column=2, value=m_d["sales"])
            ws_chart_data.cell(row=m_idx, column=3, value=m_d["count"])
            ws_chart_data.cell(row=m_idx, column=4, value=m_d["cgst"])
            ws_chart_data.cell(row=m_idx, column=5, value=m_d["sgst"])
            ws_chart_data.cell(row=m_idx, column=6, value=m_d["igst"])
            ws_chart_data.cell(row=m_idx, column=7, value=m_d["taxable"])
            m_idx += 1
        m_end_row = m_idx - 1

        # 2. Top Customers (Top 10, Sorted Descending)
        top_custs = sorted(party_map.items(), key=lambda x: x[1]["grand"], reverse=True)[:10]
        if not top_custs:
            top_custs = [("No Customers", {"grand": 0.0})]

        ws_chart_data.cell(row=1, column=9, value="Customer")
        ws_chart_data.cell(row=1, column=10, value="Sales (₹)")
        c_idx = 2
        for cname, cd in top_custs:
            ws_chart_data.cell(row=c_idx, column=9, value=cname)
            ws_chart_data.cell(row=c_idx, column=10, value=cd["grand"])
            c_idx += 1
        c_end_row = c_idx - 1

        # 3. Top Products (Top 10, Sorted Descending)
        prod_sales_map = defaultdict(float)
        for line in items_db:
            prod_sales_map[line.product_name_snapshot] += float(line.line_total)
        top_prods = sorted(prod_sales_map.items(), key=lambda x: x[1], reverse=True)[:10]
        if not top_prods:
            top_prods = [("No Products", 0.0)]

        ws_chart_data.cell(row=1, column=12, value="Product")
        ws_chart_data.cell(row=1, column=13, value="Revenue (₹)")
        p_idx = 2
        for pname, prev in top_prods:
            ws_chart_data.cell(row=p_idx, column=12, value=pname)
            ws_chart_data.cell(row=p_idx, column=13, value=prev)
            p_idx += 1
        p_end_row = p_idx - 1

        # 4. Sales by State
        state_map = defaultdict(float)
        for inv in invoices_db:
            st = inv.place_of_supply or "Unspecified"
            state_map[st] += float(inv.grand_total)
        sorted_states = sorted(state_map.items(), key=lambda x: x[1], reverse=True)
        if not sorted_states:
            sorted_states = [("Unspecified", 0.0)]

        ws_chart_data.cell(row=1, column=15, value="State")
        ws_chart_data.cell(row=1, column=16, value="Sales (₹)")
        st_idx = 2
        for st_name, st_val in sorted_states:
            ws_chart_data.cell(row=st_idx, column=15, value=st_name)
            ws_chart_data.cell(row=st_idx, column=16, value=st_val)
            st_idx += 1
        st_end_row = st_idx - 1

        # 5. CGST / SGST / IGST Distribution
        tot_cgst = sum(float(inv.cgst_total) for inv in invoices_db)
        tot_sgst = sum(float(inv.sgst_total) for inv in invoices_db)
        tot_igst = sum(float(inv.igst_total) for inv in invoices_db)
        tot_cess = sum(float(inv.cess_total) for inv in invoices_db)
        tax_dist = [("CGST", tot_cgst), ("SGST", tot_sgst), ("IGST", tot_igst), ("Cess", tot_cess)]
        non_zero_tax = [t for t in tax_dist if t[1] > 0]
        if not non_zero_tax:
            non_zero_tax = [("GST Tax", 0.0)]

        ws_chart_data.cell(row=1, column=18, value="Tax Type")
        ws_chart_data.cell(row=1, column=19, value="Amount (₹)")
        td_idx = 2
        for tname, tamt in non_zero_tax:
            ws_chart_data.cell(row=td_idx, column=18, value=tname)
            ws_chart_data.cell(row=td_idx, column=19, value=tamt)
            td_idx += 1
        td_end_row = td_idx - 1

        # 6. Invoice Status Distribution
        status_map = defaultdict(int)
        for inv in invoices_db:
            status_map[inv.status.capitalize()] += 1
        if not status_map:
            status_map["Issued"] = 0

        ws_chart_data.cell(row=1, column=21, value="Status")
        ws_chart_data.cell(row=1, column=22, value="Count")
        stat_idx = 2
        for sname, scount in status_map.items():
            ws_chart_data.cell(row=stat_idx, column=21, value=sname)
            ws_chart_data.cell(row=stat_idx, column=22, value=scount)
            stat_idx += 1
        stat_end_row = stat_idx - 1

        # 7. HSN Revenue
        hsn_rev_list = sorted(hsn_map.items(), key=lambda x: x[1]["taxable"], reverse=True)[:10]
        if not hsn_rev_list:
            hsn_rev_list = [("N/A", {"desc": "N/A", "taxable": 0.0})]

        ws_chart_data.cell(row=1, column=24, value="HSN/SAC")
        ws_chart_data.cell(row=1, column=25, value="Taxable Revenue (₹)")
        h_idx = 2
        for hsn_code, hd in hsn_rev_list:
            ws_chart_data.cell(row=h_idx, column=24, value=hsn_code)
            ws_chart_data.cell(row=h_idx, column=25, value=hd["taxable"])
            h_idx += 1
        h_end_row = h_idx - 1

        ws_chart_data.sheet_state = "hidden"

        # ---------------------------------------------------------
        # 7. Dashboard Sheet Visual Redesign
        # ---------------------------------------------------------
        ws_dash = wb.create_sheet(title="Dashboard", index=0)
        ws_dash.sheet_view.zoomScale = 85
        ws_dash.views.sheetView[0].showGridLines = False

        # Set Column Widths for Canvas (A1:N70 Grid)
        col_widths = {
            "A": 3, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14,
            "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14,
            "M": 14, "N": 3
        }
        for col_let, w in col_widths.items():
            ws_dash.column_dimensions[col_let].width = w

        # Header Title Block (Row 1 to 3)
        ws_dash.row_dimensions[1].height = 24
        ws_dash.row_dimensions[2].height = 28
        ws_dash.row_dimensions[3].height = 20

        t1 = ws_dash.cell(row=1, column=2, value="ADVANCE BILLING")
        t1.font = Font(name="Calibri", size=11, bold=True, color="FF7A00")

        t2 = ws_dash.cell(row=2, column=2, value="Business & Billing Performance Dashboard")
        t2.font = Font(name="Calibri", size=18, bold=True, color="0F172A")

        fy_str = f"FY {now.year}-{str(now.year + 1)[-2:]}" if now.month >= 4 else f"FY {now.year - 1}-{str(now.year)[-2:]}"
        t3 = ws_dash.cell(
            row=3, column=2,
            value=f"Organization: {organization.business_name}  |  Financial Year: {fy_str}  |  Generated: {now.strftime('%d %b %Y %H:%M UTC')}"
        )
        t3.font = Font(name="Calibri", size=10, italic=True, color="64748B")

        # ---------------------------------------------------------
        # KPI Cards (8 Cards in 2 Rows of 4 Cards)
        # Row 5-6 (Cards 1 to 4) & Row 8-9 (Cards 5 to 8)
        # ---------------------------------------------------------
        kpis = [
            ("Total Sales (₹)", "=SUM(Invoices!Z2:Z10000)", 5, 2, 3),        # B5:C6
            ("Total Invoices", "=COUNTA(Invoices!A2:A10000)", 5, 5, 6),       # E5:F6
            ("Customers", "=COUNTA(Customers!A2:A10000)", 5, 8, 9),          # H5:I6
            ("Products", "=COUNTA(Products!A2:A10000)", 5, 11, 12),         # K5:L6
            ("Taxable Value (₹)", "=SUM(Invoices!T2:T10000)", 8, 2, 3),      # B8:C9
            ("Total Tax (₹)", "=SUM(Invoices!U2:X10000)", 8, 5, 6),         # E8:F9
            ("Average Invoice (₹)", "=IFERROR(B6/E6, 0)", 8, 8, 9),         # H8:I9
            ("Cancelled Invoices", '=COUNTIF(Invoices!C2:C10000, "cancelled")', 8, 11, 12), # K8:L9
        ]

        card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        card_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        for label, formula, start_r, start_c, end_c in kpis:
            # Label Cell (Upper Row)
            lbl_cell = ws_dash.cell(row=start_r, column=start_c, value=label)
            lbl_cell.font = Font(name="Calibri", size=9, bold=True, color="64748B")
            lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl_cell.fill = card_fill

            # Value Cell (Lower Row)
            val_cell = ws_dash.cell(row=start_r + 1, column=start_c, value=formula)
            val_cell.font = Font(name="Calibri", size=14, bold=True, color="FF7A00" if "Sales" in label or "Tax" in label else "0F172A")
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            val_cell.fill = card_fill

            # Apply merge and border across start_c to end_c
            ws_dash.merge_cells(start_row=start_r, start_column=start_c, end_row=start_r, end_column=end_c)
            ws_dash.merge_cells(start_row=start_r + 1, start_column=start_c, end_row=start_r + 1, end_column=end_c)

            for r in range(start_r, start_r + 2):
                for c in range(start_c, end_c + 1):
                    ws_dash.cell(row=r, column=c).border = card_border
                    ws_dash.cell(row=r, column=c).fill = card_fill

        ws_dash.row_dimensions[5].height = 18
        ws_dash.row_dimensions[6].height = 26
        ws_dash.row_dimensions[8].height = 18
        ws_dash.row_dimensions[9].height = 26

        # ---------------------------------------------------------
        # 8. Create & Embed 10 Professional openpyxl Charts
        # ---------------------------------------------------------
        num_months = len(monthly_map)

        # CHART 1: Monthly Sales Trend (LineChart for 4+ months, BarChart for <=3)
        if num_months >= 4:
            c1 = LineChart()
            c1.style = 13
        else:
            c1 = BarChart()
            c1.style = 10
        c1.title = "Monthly Sales Trend"
        c1.y_axis.title = "Sales (₹)"
        c1.x_axis.title = "Month"
        c1.legend = None  # Remove redundant legend for single series
        data1 = Reference(ws_chart_data, min_col=2, min_row=1, max_row=m_end_row)
        cats1 = Reference(ws_chart_data, min_col=1, min_row=2, max_row=m_end_row)
        c1.add_data(data1, titles_from_data=True)
        c1.set_categories(cats1)
        c1.width = 15.5
        c1.height = 8.5
        ws_dash.add_chart(c1, "B12")

        # CHART 2: Monthly Invoice Count (BarChart)
        c2 = BarChart()
        c2.title = "Monthly Invoice Count"
        c2.style = 10
        c2.y_axis.title = "Invoices"
        c2.x_axis.title = "Month"
        c2.legend = None
        data2 = Reference(ws_chart_data, min_col=3, min_row=1, max_row=m_end_row)
        c2.add_data(data2, titles_from_data=True)
        c2.set_categories(cats1)
        c2.width = 15.5
        c2.height = 8.5
        ws_dash.add_chart(c2, "H12")

        # CHART 3: Monthly GST Trend (Multi-Series LineChart -> KEEP LEGEND)
        c3 = LineChart()
        c3.title = "Monthly GST Tax Breakdown"
        c3.style = 13
        c3.y_axis.title = "Tax (₹)"
        c3.x_axis.title = "Month"
        data3 = Reference(ws_chart_data, min_col=4, min_row=1, max_col=6, max_row=m_end_row)
        c3.add_data(data3, titles_from_data=True)
        c3.set_categories(cats1)
        c3.width = 15.5
        c3.height = 8.5
        ws_dash.add_chart(c3, "B28")

        # CHART 4: Monthly Taxable Value (BarChart)
        c4 = BarChart()
        c4.title = "Monthly Taxable Value"
        c4.style = 11
        c4.y_axis.title = "Taxable Value (₹)"
        c4.x_axis.title = "Month"
        c4.legend = None
        data4 = Reference(ws_chart_data, min_col=7, min_row=1, max_row=m_end_row)
        c4.add_data(data4, titles_from_data=True)
        c4.set_categories(cats1)
        c4.width = 15.5
        c4.height = 8.5
        ws_dash.add_chart(c4, "H28")

        # CHART 5: Top 10 Customers by Sales (Horizontal BarChart)
        c5 = BarChart()
        c5.type = "bar"
        c5.title = "Top 10 Customers by Sales"
        c5.style = 12
        c5.x_axis.title = "Sales (₹)"
        c5.legend = None
        data5 = Reference(ws_chart_data, min_col=10, min_row=1, max_row=c_end_row)
        cats5 = Reference(ws_chart_data, min_col=9, min_row=2, max_row=c_end_row)
        c5.add_data(data5, titles_from_data=True)
        c5.set_categories(cats5)
        c5.width = 15.5
        c5.height = 8.5
        ws_dash.add_chart(c5, "B44")

        # CHART 6: Top 10 Products by Revenue (Horizontal BarChart)
        c6 = BarChart()
        c6.type = "bar"
        c6.title = "Top 10 Products by Revenue"
        c6.style = 13
        c6.x_axis.title = "Revenue (₹)"
        c6.legend = None
        data6 = Reference(ws_chart_data, min_col=13, min_row=1, max_row=p_end_row)
        cats6 = Reference(ws_chart_data, min_col=12, min_row=2, max_row=p_end_row)
        c6.add_data(data6, titles_from_data=True)
        c6.set_categories(cats6)
        c6.width = 15.5
        c6.height = 8.5
        ws_dash.add_chart(c6, "H44")

        # CHART 7: Sales by State (Horizontal BarChart)
        c7 = BarChart()
        c7.type = "bar"
        c7.title = "Sales by State / POS"
        c7.style = 10
        c7.x_axis.title = "Sales (₹)"
        c7.legend = None
        data7 = Reference(ws_chart_data, min_col=16, min_row=1, max_row=st_end_row)
        cats7 = Reference(ws_chart_data, min_col=15, min_row=2, max_row=st_end_row)
        c7.add_data(data7, titles_from_data=True)
        c7.set_categories(cats7)
        c7.width = 15.5
        c7.height = 8.5
        ws_dash.add_chart(c7, "B60")

        # CHART 8: CGST / SGST / IGST Distribution (DoughnutChart)
        c8 = DoughnutChart()
        c8.title = "GST Tax Type Distribution"
        data8 = Reference(ws_chart_data, min_col=19, min_row=1, max_row=td_end_row)
        cats8 = Reference(ws_chart_data, min_col=18, min_row=2, max_row=td_end_row)
        c8.add_data(data8, titles_from_data=True)
        c8.set_categories(cats8)
        c8.width = 15.5
        c8.height = 8.5
        ws_dash.add_chart(c8, "H60")

        # CHART 9: Invoice Status Distribution (DoughnutChart)
        c9 = DoughnutChart()
        c9.title = "Invoice Status Breakdown"
        data9 = Reference(ws_chart_data, min_col=22, min_row=1, max_row=stat_end_row)
        cats9 = Reference(ws_chart_data, min_col=21, min_row=2, max_row=stat_end_row)
        c9.add_data(data9, titles_from_data=True)
        c9.set_categories(cats9)
        c9.width = 15.5
        c9.height = 8.5
        ws_dash.add_chart(c9, "B76")

        # CHART 10: HSN/SAC Revenue (Horizontal BarChart)
        c10 = BarChart()
        c10.type = "bar"
        c10.title = "Top 10 HSN/SAC Codes by Revenue"
        c10.style = 12
        c10.x_axis.title = "Taxable Revenue (₹)"
        c10.legend = None
        data10 = Reference(ws_chart_data, min_col=25, min_row=1, max_row=h_end_row)
        cats10 = Reference(ws_chart_data, min_col=24, min_row=2, max_row=h_end_row)
        c10.add_data(data10, titles_from_data=True)
        c10.set_categories(cats10)
        c10.width = 15.5
        c10.height = 8.5
        ws_dash.add_chart(c10, "H76")

        # Clean unused default sheet
        if default_ws and default_ws.title in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(default_ws)

        # Output Buffer
        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()
        buf.close()

        record_counts = {
            "organization": len(org_rows),
            "customers": len(cust_rows),
            "customer_addresses": len(addr_rows),
            "products": len(prod_rows),
            "invoices": len(inv_rows),
            "invoice_items": len(item_rows),
        }

        manifest = {
            "signature": SIGNATURE,
            "schema_version": SCHEMA_VERSION,
            "organization_id": organization.id,
            "organization_name": organization.business_name,
            "created_at": now.isoformat(),
            "record_counts": record_counts,
            "total_records": sum(record_counts.values()),
        }

        return excel_bytes, filename, manifest
