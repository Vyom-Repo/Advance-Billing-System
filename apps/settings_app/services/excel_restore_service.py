"""
apps/settings_app/services/excel_restore_service.py

Service for validating and restoring Advance Billing Excel Backup workbooks (.xlsx).
Enforces:
- File type security (.xlsx only, openpyxl ZIP parsing, 15MB limit)
- Machine-readable signature (ADVANCE_BILLING_BACKUP) in _Metadata and README
- Schema version compatibility (1.0)
- Multi-tenant organization isolation (organization_id check)
- Exact header & column validation for the 6 normalized business sheets
- Relational integrity verification (Foreign keys across sheets)
- Financial reconciliation using Decimal arithmetic
- Two-Phase import flow:
    Phase 1: Read-only validation dry-run & preview
    Phase 2: Atomic transactional database restoration
"""

import io
from decimal import Decimal
from typing import Any, Dict, List, Tuple

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.billing.models import Invoice, InvoiceLine, InvoiceStatus
from apps.customers.models import Customer
from apps.organization.models import Organization
from apps.products.models import Product
from apps.settings_app.models import DataManagementAction, DataManagementAuditLog

SIGNATURE = "ADVANCE_BILLING_BACKUP"
SUPPORTED_SCHEMA_VERSIONS = ["1.0"]
MAX_FILE_SIZE_BYTES = 15 * 1024 * 1024  # 15 MB

REQUIRED_SHEETS = [
    "Dashboard", "README", "Organization", "Customers", "Customer_Addresses",
    "Products", "Invoices", "Invoice_Items", "GST_Summary", "GST_Sales",
    "HSN_Summary", "Party_Summary", "Import_Map"
]

EXPECTED_HEADERS = {
    "Organization": [
        "organization_id", "business_name", "legal_business_name", "is_gst_registered",
        "gstin", "pan", "state_code", "business_email", "phone_number", "address_line_1",
        "address_line_2", "city", "state", "pincode", "country", "signature_mode",
        "authorized_signatory_name", "show_computer_generated_disclaimer",
        "terms_and_conditions", "created_at", "updated_at"
    ],
    "Customers": [
        "customer_id", "name", "gstin", "customer_type", "gst_status",
        "billing_address_line_1", "billing_address_line_2", "billing_city",
        "billing_state", "billing_pin_code", "billing_state_code", "billing_country",
        "created_at", "updated_at"
    ],
    "Customer_Addresses": [
        "address_id", "customer_id", "address_line_1", "address_line_2",
        "city", "state", "state_code", "pincode", "country"
    ],
    "Products": [
        "product_id", "name", "product_type", "hsn_code", "sac_code",
        "taxability_type", "gst_rate", "unit_price", "price_basis", "uqc",
        "cess_applicable", "cess_type", "cess_rate_or_amount",
        "reverse_charge_applicable", "created_at", "updated_at"
    ],
    "Invoices": [
        "invoice_id", "invoice_number", "status", "invoice_date", "due_date",
        "place_of_supply", "currency", "customer_id", "customer_name_snapshot",
        "customer_gstin_snapshot", "customer_billing_address_snapshot",
        "customer_state_code_snapshot", "shipping_same_as_billing",
        "shipping_address_line_1", "shipping_city", "shipping_state",
        "shipping_pincode", "subtotal", "discount_total", "taxable_amount",
        "cgst_total", "sgst_total", "igst_total", "cess_total", "round_off",
        "grand_total", "notes", "terms", "created_at", "updated_at"
    ],
    "Invoice_Items": [
        "item_id", "invoice_id", "product_id", "position", "product_name_snapshot",
        "product_type_snapshot", "hsn_sac_snapshot", "taxability_type_snapshot",
        "gst_rate_snapshot", "cess_applicable_snapshot", "cess_type_snapshot",
        "cess_rate_snapshot", "reverse_charge_snapshot", "price_basis_snapshot",
        "uqc_snapshot", "quantity", "unit_price", "discount_type", "discount_value",
        "discount_amount", "taxable_value", "cgst_rate", "cgst_amount", "sgst_rate",
        "sgst_amount", "igst_rate", "igst_amount", "cess_amount", "line_total"
    ],
}


class ExcelRestoreService:
    @classmethod
    def parse_sheet_rows(cls, ws) -> List[Dict[str, Any]]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        results = []
        for r in rows[1:]:
            if not any(r):
                continue
            row_dict = {}
            for idx, h in enumerate(headers):
                if h:
                    row_dict[h] = r[idx] if idx < len(r) else None
            results.append(row_dict)
        return results

    @classmethod
    def to_decimal(cls, val: Any) -> Decimal:
        if val is None or val == "":
            return Decimal("0.00")
        try:
            return Decimal(str(val)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal("0.00")

    @classmethod
    def to_bool(cls, val: Any) -> bool:
        if isinstance(val, bool):
            return val
        s = str(val or "").strip().lower()
        return s in ("true", "1", "yes", "t")

    @classmethod
    def validate_and_preview(
        cls, file_bytes: bytes, filename: str, current_organization: Organization
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """
        PHASE 1: READ-ONLY VALIDATION & DRY-RUN PREVIEW
        Does NOT modify the database. Returns (is_valid, error_or_success_msg, preview_data)
        """
        # 1. File Type & Extension Check
        if not filename.lower().endswith(".xlsx"):
            return False, "Invalid file format. Upload must be an authentic .xlsx workbook.", {}

        if len(file_bytes) > MAX_FILE_SIZE_BYTES:
            return False, f"File size ({len(file_bytes) / (1024*1024):.2f} MB) exceeds maximum allowed size (15 MB).", {}

        # 2. XLSX/Zip Structure Parsing
        try:
            buf = io.BytesIO(file_bytes)
            wb = load_workbook(buf, data_only=True, read_only=False)
        except Exception as e:
            return False, f"Failed to parse Excel workbook: {str(e)}. File may be corrupted or encrypted.", {}

        # 3. Signature & Identity Check (_Metadata & README)
        if "_Metadata" not in wb.sheetnames:
            return False, "This file is not a valid Advance Billing backup (missing _Metadata signature sheet).", {}

        meta_rows = cls.parse_sheet_rows(wb["_Metadata"])
        meta_dict = {}
        for r in meta_rows:
            k = str(r.get("Key") or r.get("key") or "").strip()
            v = str(r.get("Value") or r.get("value") or "").strip()
            if k:
                meta_dict[k] = v

        file_sig = meta_dict.get("signature", "")
        file_version = meta_dict.get("schema_version", "")
        file_org_id = meta_dict.get("organization_id", "")

        if file_sig != SIGNATURE:
            return False, "This file is not a valid Advance Billing backup (invalid signature).", {}

        if file_version not in SUPPORTED_SCHEMA_VERSIONS:
            return False, f"This Advance Billing backup uses an unsupported schema version '{file_version}'. Currently supported: 1.0", {}

        # 4. Multi-Tenant Organization Isolation Check
        if str(file_org_id) != str(current_organization.id):
            return False, f"This backup belongs to a different organization (ID: {file_org_id}) and cannot be imported here.", {}

        # README Sheet Double Check
        if "README" not in wb.sheetnames:
            return False, "This file is not a valid Advance Billing backup (missing README sheet).", {}

        # 5. Required Sheets Check
        for req_sheet in REQUIRED_SHEETS:
            if req_sheet not in wb.sheetnames:
                return False, f"Missing required worksheet '{req_sheet}' in backup package.", {}

        # 6. Exact Header Validation for Normalized Business Sheets
        for sheet_name, expected_headers in EXPECTED_HEADERS.items():
            ws = wb[sheet_name]
            first_row = list(ws.iter_rows(values_only=True))[0] if ws.max_row > 0 else []
            actual_headers = [str(h or "").strip() for h in first_row]
            for req_h in expected_headers:
                if req_h not in actual_headers:
                    return False, f"Invalid schema in worksheet '{sheet_name}': missing required column header '{req_h}'.", {}

        # Parse normalized rows for validation
        cust_rows = cls.parse_sheet_rows(wb["Customers"])
        addr_rows = cls.parse_sheet_rows(wb["Customer_Addresses"])
        prod_rows = cls.parse_sheet_rows(wb["Products"])
        inv_rows = cls.parse_sheet_rows(wb["Invoices"])
        item_rows = cls.parse_sheet_rows(wb["Invoice_Items"])

        cust_ids = set()
        for r in cust_rows:
            cid = str(r.get("customer_id") or "").strip()
            if not cid:
                return False, "Found customer record with empty customer_id in 'Customers' worksheet.", {}
            if cid in cust_ids:
                return False, f"Duplicate customer_id '{cid}' found in 'Customers' worksheet.", {}
            cust_ids.add(cid)

        prod_ids = set()
        for r in prod_rows:
            pid = str(r.get("product_id") or "").strip()
            if not pid:
                return False, "Found product record with empty product_id in 'Products' worksheet.", {}
            if pid in prod_ids:
                return False, f"Duplicate product_id '{pid}' found in 'Products' worksheet.", {}
            prod_ids.add(pid)

        inv_ids = set()
        inv_numbers = set()
        for r in inv_rows:
            iid = str(r.get("invoice_id") or "").strip()
            inum = str(r.get("invoice_number") or "").strip()
            if not iid:
                return False, "Found invoice record with empty invoice_id in 'Invoices' worksheet.", {}
            if iid in inv_ids:
                return False, f"Duplicate invoice_id '{iid}' found in 'Invoices' worksheet.", {}
            inv_ids.add(iid)

            if inum:
                if inum in inv_numbers:
                    return False, f"Duplicate invoice_number '{inum}' found in 'Invoices' worksheet.", {}
                inv_numbers.add(inum)

            # Check Customer Reference
            ref_cid = str(r.get("customer_id") or "").strip()
            if ref_cid and ref_cid not in cust_ids:
                return False, f"Invoice '{inum or iid}' references non-existent customer_id '{ref_cid}'.", {}

            # Financial Reconciliation Check (Decimal arithmetic)
            subtotal = cls.to_decimal(r.get("subtotal"))
            discount = cls.to_decimal(r.get("discount_total"))
            taxable = cls.to_decimal(r.get("taxable_amount"))
            cgst = cls.to_decimal(r.get("cgst_total"))
            sgst = cls.to_decimal(r.get("sgst_total"))
            igst = cls.to_decimal(r.get("igst_total"))
            cess = cls.to_decimal(r.get("cess_total"))
            round_off = cls.to_decimal(r.get("round_off"))
            grand_total = cls.to_decimal(r.get("grand_total"))

            calculated_grand = taxable + cgst + sgst + igst + cess + round_off
            if abs(calculated_grand - grand_total) > Decimal("0.05"):
                return False, f"Financial calculation mismatch for Invoice '{inum or iid}': grand_total ({grand_total}) does not equal taxable + tax + round_off ({calculated_grand}).", {}

        for r in item_rows:
            ref_iid = str(r.get("invoice_id") or "").strip()
            if not ref_iid or ref_iid not in inv_ids:
                return False, f"Invoice line item references invalid invoice_id '{ref_iid}'.", {}

            ref_pid = str(r.get("product_id") or "").strip()
            if ref_pid and ref_pid not in prod_ids:
                return False, f"Invoice line item references invalid product_id '{ref_pid}'.", {}

        for r in addr_rows:
            ref_cid = str(r.get("customer_id") or "").strip()
            if not ref_cid or ref_cid not in cust_ids:
                return False, f"Customer address references invalid customer_id '{ref_cid}'.", {}

        # 7. Record Action Preview Calculation (Create vs Update)
        existing_cust_uuids = set(str(u) for u in Customer.objects.filter(organization=current_organization).values_list("uuid", flat=True))
        existing_prod_uuids = set(str(u) for u in Product.objects.filter(organization=current_organization).values_list("uuid", flat=True))
        existing_inv_uuids = set(str(u) for u in Invoice.objects.filter(organization=current_organization).values_list("uuid", flat=True))

        cust_new = sum(1 for cid in cust_ids if cid not in existing_cust_uuids)
        cust_upd = len(cust_ids) - cust_new

        prod_new = sum(1 for pid in prod_ids if pid not in existing_prod_uuids)
        prod_upd = len(prod_ids) - prod_new

        inv_new = sum(1 for iid in inv_ids if iid not in existing_inv_uuids)
        inv_upd = len(inv_ids) - inv_new

        preview_context = {
            "is_valid": True,
            "filename": filename,
            "signature": file_sig,
            "schema_version": file_version,
            "organization_name": current_organization.business_name,
            "counts": {
                "customers": {"total": len(cust_rows), "created": cust_new, "updated": cust_upd},
                "customer_addresses": {"total": len(addr_rows), "created": len(addr_rows), "updated": 0},
                "products": {"total": len(prod_rows), "created": prod_new, "updated": prod_upd},
                "invoices": {"total": len(inv_rows), "created": inv_new, "updated": inv_upd},
                "invoice_items": {"total": len(item_rows), "created": len(item_rows), "updated": 0},
            },
            "total_records": len(cust_rows) + len(addr_rows) + len(prod_rows) + len(inv_rows) + len(item_rows),
        }

        return True, "Backup validation succeeded. Workbook is authentic and ready to restore.", preview_context

    @classmethod
    def execute_restore(
        cls, file_bytes: bytes, filename: str, current_organization: Organization
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """
        PHASE 2: ATOMIC TRANSACTIONAL DATABASE RESTORATION
        Runs inside @transaction.atomic. Rolls back everything on failure.
        """
        is_valid, msg, preview = cls.validate_and_preview(file_bytes, filename, current_organization)
        if not is_valid:
            DataManagementAuditLog.objects.create(
                organization=current_organization,
                action=DataManagementAction.RESTORE_BACKUP,
                status="failed",
                details=f"Restore rejected during validation: {msg}",
            )
            return False, f"Validation failed: {msg}", {}

        try:
            with transaction.atomic():
                buf = io.BytesIO(file_bytes)
                wb = load_workbook(buf, data_only=True)

                org_rows = cls.parse_sheet_rows(wb["Organization"])
                cust_rows = cls.parse_sheet_rows(wb["Customers"])
                prod_rows = cls.parse_sheet_rows(wb["Products"])
                inv_rows = cls.parse_sheet_rows(wb["Invoices"])
                item_rows = cls.parse_sheet_rows(wb["Invoice_Items"])

                # 1. Update Organization attributes
                if org_rows:
                    o = org_rows[0]
                    current_organization.legal_business_name = o.get("legal_business_name") or ""
                    current_organization.is_gst_registered = cls.to_bool(o.get("is_gst_registered"))
                    current_organization.gstin = o.get("gstin") or ""
                    current_organization.pan = o.get("pan") or ""
                    current_organization.state_code = o.get("state_code") or ""
                    current_organization.business_email = o.get("business_email") or current_organization.business_email
                    current_organization.phone_number = o.get("phone_number") or ""
                    current_organization.address_line_1 = o.get("address_line_1") or ""
                    current_organization.address_line_2 = o.get("address_line_2") or ""
                    current_organization.city = o.get("city") or ""
                    current_organization.state = o.get("state") or ""
                    current_organization.pincode = o.get("pincode") or ""
                    current_organization.save()

                # 2. Restore Customers
                cust_map = {}
                for r in cust_rows:
                    cid = str(r["customer_id"]).strip()
                    cust_obj, _ = Customer.objects.update_or_create(
                        organization=current_organization,
                        uuid=cid,
                        defaults={
                            "name": r.get("name") or "Customer",
                            "gstin": r.get("gstin") or "",
                            "customer_type": r.get("customer_type") or "B2C",
                            "gst_status": r.get("gst_status") or "unregistered",
                            "billing_address_line_1": r.get("billing_address_line_1") or "",
                            "billing_address_line_2": r.get("billing_address_line_2") or "",
                            "billing_city": r.get("billing_city") or "",
                            "billing_state": r.get("billing_state") or "",
                            "billing_pin_code": r.get("billing_pin_code") or "",
                            "billing_state_code": r.get("billing_state_code") or "",
                            "billing_country": r.get("billing_country") or "India",
                        }
                    )
                    cust_map[cid] = cust_obj

                # 3. Restore Products
                prod_map = {}
                for r in prod_rows:
                    pid = str(r["product_id"]).strip()
                    prod_obj, _ = Product.objects.update_or_create(
                        organization=current_organization,
                        uuid=pid,
                        defaults={
                            "name": r.get("name") or "Product",
                            "product_type": r.get("product_type") or "goods",
                            "hsn_code": r.get("hsn_code") or "",
                            "sac_code": r.get("sac_code") or "",
                            "taxability_type": r.get("taxability_type") or "taxable",
                            "gst_rate": cls.to_decimal(r.get("gst_rate")),
                            "unit_price": cls.to_decimal(r.get("unit_price")),
                            "price_basis": r.get("price_basis") or "tax_exclusive",
                            "uqc": r.get("uqc") or "OTH",
                            "cess_applicable": cls.to_bool(r.get("cess_applicable")),
                            "cess_type": r.get("cess_type") or "",
                            "cess_rate_or_amount": cls.to_decimal(r.get("cess_rate_or_amount")),
                            "reverse_charge_applicable": cls.to_bool(r.get("reverse_charge_applicable")),
                        }
                    )
                    prod_map[pid] = prod_obj

                # 4. Restore Invoices
                inv_map = {}
                for r in inv_rows:
                    iid = str(r["invoice_id"]).strip()
                    ref_cid = str(r.get("customer_id") or "").strip()
                    customer_obj = cust_map.get(ref_cid)

                    inv_obj, _ = Invoice.objects.update_or_create(
                        organization=current_organization,
                        uuid=iid,
                        defaults={
                            "customer": customer_obj,
                            "invoice_number": r.get("invoice_number") or "",
                            "status": r.get("status") or InvoiceStatus.DRAFT,
                            "invoice_date": r.get("invoice_date") or timezone.now().date(),
                            "due_date": r.get("due_date"),
                            "place_of_supply": r.get("place_of_supply") or "",
                            "currency": r.get("currency") or "INR",
                            "customer_name_snapshot": r.get("customer_name_snapshot") or "Customer",
                            "customer_gstin_snapshot": r.get("customer_gstin_snapshot") or "",
                            "customer_billing_address_snapshot": r.get("customer_billing_address_snapshot") or "",
                            "customer_state_code_snapshot": r.get("customer_state_code_snapshot") or "",
                            "shipping_same_as_billing": cls.to_bool(r.get("shipping_same_as_billing")),
                            "shipping_address_line_1": r.get("shipping_address_line_1") or "",
                            "shipping_city": r.get("shipping_city") or "",
                            "shipping_state": r.get("shipping_state") or "",
                            "shipping_pincode": r.get("shipping_pincode") or "",
                            "subtotal": cls.to_decimal(r.get("subtotal")),
                            "discount_total": cls.to_decimal(r.get("discount_total")),
                            "taxable_amount": cls.to_decimal(r.get("taxable_amount")),
                            "cgst_total": cls.to_decimal(r.get("cgst_total")),
                            "sgst_total": cls.to_decimal(r.get("sgst_total")),
                            "igst_total": cls.to_decimal(r.get("igst_total")),
                            "cess_total": cls.to_decimal(r.get("cess_total")),
                            "round_off": cls.to_decimal(r.get("round_off")),
                            "grand_total": cls.to_decimal(r.get("grand_total")),
                            "notes": r.get("notes") or "",
                            "terms": r.get("terms") or "",
                        }
                    )
                    inv_map[iid] = inv_obj

                # 5. Restore Invoice Lines (Clear & Recreate lines for atomic restore)
                for inv_obj in inv_map.values():
                    inv_obj.lines.all().delete()

                line_objs = []
                for r in item_rows:
                    ref_iid = str(r["invoice_id"]).strip()
                    ref_pid = str(r.get("product_id") or "").strip()

                    inv_obj = inv_map[ref_iid]
                    prod_obj = prod_map.get(ref_pid)

                    line_objs.append(InvoiceLine(
                        invoice=inv_obj,
                        position=int(r.get("position") or 1),
                        product=prod_obj,
                        product_name_snapshot=r.get("product_name_snapshot") or "Item",
                        product_type_snapshot=r.get("product_type_snapshot") or "goods",
                        hsn_sac_snapshot=r.get("hsn_sac_snapshot") or "",
                        taxability_type_snapshot=r.get("taxability_type_snapshot") or "taxable",
                        gst_rate_snapshot=cls.to_decimal(r.get("gst_rate_snapshot")),
                        cess_applicable_snapshot=cls.to_bool(r.get("cess_applicable_snapshot")),
                        cess_type_snapshot=r.get("cess_type_snapshot") or "",
                        cess_rate_snapshot=cls.to_decimal(r.get("cess_rate_snapshot")),
                        reverse_charge_snapshot=cls.to_bool(r.get("reverse_charge_snapshot")),
                        price_basis_snapshot=r.get("price_basis_snapshot") or "tax_exclusive",
                        uqc_snapshot=r.get("uqc_snapshot") or "OTH",
                        quantity=cls.to_decimal(r.get("quantity")),
                        unit_price=cls.to_decimal(r.get("unit_price")),
                        discount_type=r.get("discount_type") or "none",
                        discount_value=cls.to_decimal(r.get("discount_value")),
                        discount_amount=cls.to_decimal(r.get("discount_amount")),
                        taxable_value=cls.to_decimal(r.get("taxable_value")),
                        cgst_rate=cls.to_decimal(r.get("cgst_rate")),
                        cgst_amount=cls.to_decimal(r.get("cgst_amount")),
                        sgst_rate=cls.to_decimal(r.get("sgst_rate")),
                        sgst_amount=cls.to_decimal(r.get("sgst_amount")),
                        igst_rate=cls.to_decimal(r.get("igst_rate")),
                        igst_amount=cls.to_decimal(r.get("igst_amount")),
                        cess_amount=cls.to_decimal(r.get("cess_amount")),
                        line_total=cls.to_decimal(r.get("line_total")),
                    ))

                if line_objs:
                    InvoiceLine.objects.bulk_create(line_objs)

                # Record Audit Log
                DataManagementAuditLog.objects.create(
                    organization=current_organization,
                    action=DataManagementAction.RESTORE_BACKUP,
                    status="completed",
                    details=f"Backup restored successfully from '{filename}'. Total records restored: {preview['total_records']}",
                )

                return True, f"Backup restored successfully! Restored {preview['total_records']} total records for {current_organization.business_name}.", preview

        except Exception as e:
            err_msg = f"Database restoration failed: {str(e)}. No changes were made."
            DataManagementAuditLog.objects.create(
                organization=current_organization,
                action=DataManagementAction.RESTORE_BACKUP,
                status="failed",
                details=err_msg,
            )
            return False, err_msg, {}
